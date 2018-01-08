# -*- coding: latin-1 -*-
# probleme sur les accents et les special_char
# # https://www.python.org/dev/peps/pep-0263/

# default value of HKEY_CURRENT_USE\Software\Microsoft\Internet Explorer\Main\Start Page


import os
import re
import sys
import getopt
import logging
import threading
import time
import datetime
from datetime import date
from msvcrt import getch
import urllib
import fileinput
import shutil
from _winreg import *
import csv
import ctypes
import subprocess
from PyQt4.QtCore import *
from PyQt4.QtGui import *
from ftplib import FTP
from ConfigParser import SafeConfigParser
import ConfigParser as cfgparser
import pysftp


path_prg = 'E:\\DISK_D\\mamitiana\\kandra\\do_not_erase\\our_tools\\'



parser = SafeConfigParser()
parser.read(path_prg + 'all_confs.txt')

# path_sublime2 = "C:\Program Files\Sublime Text 2\sublime_text.exe"
path_sublime2 = parser.get('general', 'path_subl_2')

try:
    import psycopg2,psycopg2.extras
except Exception:
    print "psycopg2 doit etre installee"
    raw_input()
    os.system("pip install psycopg2")
try:
    import xlwt
except Exception:
    print "xlwt doit etre installee"
    raw_input()
    os.system("pip install xlwt")
try:
    import xlrd
except Exception:
    print "xlrd doit etre installee"
    raw_input()
    os.system("pip install xlrd")
try:
    import xlsxwriter
except Exception:
    print "xlsxwriter doit etre installee"
    raw_input()
    os.system("pip install xlsxwriter")

try:
    import selenium
    from selenium import webdriver
    from selenium.common.exceptions import NoSuchElementException
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support.ui import Select


except Exception:
    print "selenium doit etre installee"
    raw_input()
    os.system("pip install selenium")

try:
    from PIL import Image
    pass
except Exception:
    print "PIL n_est pas installee dans cette machine"
    print "Je vous prie d_installer PIL"
    pass

try:
    import wmi
    pass
except Exception:
    os.system("pip install WMI")




# importation_installation()

reload(sys)
sys.setdefaultencoding("cp1252")


# based on: 
# # 
# # 
# # when crawling for file or for folder
# # # https://stackoverflow.com/questions/2212643/python-recursive-folder-read
# # #
# # # when checking for the depth of the crawl
# # # # https://stackoverflow.com/questions/42720627/python-os-walk-to-certain-level

"""
Colors text in console mode application (win32).
Uses ctypes and Win32 methods SetConsoleTextAttribute and
GetConsoleScreenBufferInfo.

$Id: color_console.py 534 2009-05-10 04:00:59Z andre $
"""
chrome_driver_path = 'e:\chromedriver.exe'

from ctypes import windll, Structure, c_short, c_ushort, byref

SHORT = c_short
WORD = c_ushort

class COORD(Structure):
  """struct in wincon.h."""
  # https://www.burgaud.com/bring-colors-to-the-windows-console-with-python/
  _fields_ = [
    ("X", SHORT),
    ("Y", SHORT)]

class SMALL_RECT(Structure):
  """struct in wincon.h."""
  _fields_ = [
    ("Left", SHORT),
    ("Top", SHORT),
    ("Right", SHORT),
    ("Bottom", SHORT)]

class CONSOLE_SCREEN_BUFFER_INFO(Structure):
  """struct in wincon.h."""
  _fields_ = [
    ("dwSize", COORD),
    ("dwCursorPosition", COORD),
    ("wAttributes", WORD),
    ("srWindow", SMALL_RECT),
    ("dwMaximumWindowSize", COORD)]

# winbase.h
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
STD_ERROR_HANDLE = -12

# wincon.h
FOREGROUND_BLACK     = 0x0000
FOREGROUND_BLUE      = 0x0001
FOREGROUND_GREEN     = 0x0002
FOREGROUND_CYAN      = 0x0003
FOREGROUND_RED       = 0x0004
FOREGROUND_MAGENTA   = 0x0005
FOREGROUND_YELLOW    = 0x0006
FOREGROUND_GREY      = 0x0007
FOREGROUND_INTENSITY = 0x0008 # foreground color is intensified.

BACKGROUND_BLACK     = 0x0000
BACKGROUND_BLUE      = 0x0010
BACKGROUND_GREEN     = 0x0020
BACKGROUND_CYAN      = 0x0030
BACKGROUND_RED       = 0x0040
BACKGROUND_MAGENTA   = 0x0050
BACKGROUND_YELLOW    = 0x0060
BACKGROUND_GREY      = 0x0070
BACKGROUND_INTENSITY = 0x0080 # background color is intensified.

stdout_handle = windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
SetConsoleTextAttribute = windll.kernel32.SetConsoleTextAttribute
GetConsoleScreenBufferInfo = windll.kernel32.GetConsoleScreenBufferInfo

def get_text_attr():
  """Returns the character attributes (colors) of the console screen
  buffer."""
  csbi = CONSOLE_SCREEN_BUFFER_INFO()
  GetConsoleScreenBufferInfo(stdout_handle, byref(csbi))
  return csbi.wAttributes

def set_text_attr(color):
  """Sets the character attributes (colors) of the console screen
  buffer. Color is a combination of foreground and background color,
  foreground and background intensity."""
  SetConsoleTextAttribute(stdout_handle, color)















class Person:
    def __init__(self):
        print "this is a test"











class CMail(object):

    def __init__(self, bool01 = True, 
            smtp_link = "192.168.10.4",
            tto = ["Mamitiana <mamitiana_iam@vivetic.mg>"],
            cc = [],
            sender = "mamitiana_iam@vivetic.mg",
            sujet="", 
            sbodymail="",
            file="",
            path=""):

        self.SmtpLink = smtplib.SMTP(smtp_link)
        self.html = ""
        # self.tto = ["hasina@vivetic.mg"]
        # self.cc   = ["hasina@vivetic.mg", "hasina@vivetic.mg"]
        self.tto = tto
        #self.cc   = ["Tojo <tojo_iam@vivetic.mg>","Herve <herve_iam@vivetic.mg>","Tahiry <harisoa_iam@vivetic.mg>","Fabrice <fabrice_iam@vivetic.mg>"]
        
        self.cc = cc
        # self.sender = "infodev@vivetic.mg"
        self.sender = "mamitiana_iam@vivetic.mg"
        
        self.sujet = sujet
        
        self.Email = None
        self.SmtpLink = None
        self.subject=sujet
        self.sbodymail=sbodymail
        # path = "E:/herve/Commande/CALL/CallWAV/"

        self.filename = file#"CPE_RECUP_Mail_06-05-2016.log"#
        self.attachment = open(path+file,"rb")#, "rb")#
        
        self.setUp()

    def setUp(self):
        self.Email = MIMEMultipart()
        self.Email['Date']           = time.strftime('%m/%d/%Y')
        self.Email['From']           = self.sender
        self.Email['To']      = ', '.join(self.tto)
        self.Email['Cc']      = ', '.join(self.cc)
        
        # self.SmtpLink = smtplib.SMTP("192.168.10.4")
        #self.SmtpLink = smtplib.SMTP("192.168.10.4")
        self.contentmail()

    def contentmail(self):
        html = ""
        html += """\
                <html>"""
        
        html += """\
     
        
        
        <BODY bgColor=#ffffff> 
        
        """
        
      
        html += """<DIV><FONT face=Arial size=2>&nbsp;</FONT></DIV>"""
        html += """<DIV><FONT face=Arial size=2>Bonjour,</FONT></DIV>"""
        html += """<DIV><FONT face=Arial size=2>&nbsp;</FONT></DIV>"""
        html += """<DIV><FONT face=Arial size=2>"""+self.sbodymail+""" </FONT></DIV>."""
       
        html += """<DIV><FONT face=Arial size=2>Cordialement</FONT></DIV>"""
        html += """<DIV><FONT face=Arial size=2></FONT>L'&eacute;quipe IAM</DIV>"""
        
        
        html += """\
        </body>
        </html>"""
        
        
        self.html = html.encode("utf8")
        
        
    def envoyer_mail(self):
        #try:
        attachment = self.attachment
        self.part = MIMEBase('application', 'octet-stream')
        self.part.set_payload((attachment).read())
        encoders.encode_base64(self.part)
        self.part.add_header('Content-Disposition', "attachment; filename= %s" % self.filename)
         
        self.Email.attach(self.part)
        
        self.Email.attach(
            MIMEText(
                self.html.encode(
                    'utf8','ignore'
                ), 'html'
            )
        )
        self.Email['Subject']        = self.subject
        self.SmtpLink.sendmail(
            self.Email['From'], 
            self.tto+self.cc, 
            self.Email.as_string()
        )
        self.SmtpLink.quit()
        return True
        #




class Thread001(threading.Thread): # done for Mahaitia_Demand
    

    def set_configuration(self,
        exe = "out.xlsx",
        temp_attendre = 10
    ):
        self.exe = exe
        self.temp_attendre = temp_attendre

        pass

    # class Thread001
    def __init__(
        self
    ):
        threading.Thread.__init__(self)
        fichier_conf_mahaitia = "conf_mahaitia.txt"
                    
        parser_mahaitia = SafeConfigParser()
        parser_mahaitia.read(fichier_conf_mahaitia)                    
        exe = parser_mahaitia.get('thread_conf', 'exe')
        temp_attendre = parser_mahaitia.get('thread_conf', 'temp_attendre')

        self.set_configuration(
            exe = exe,
            temp_attendre = temp_attendre
        )
        pass
    # class Thread001 #Thread001_run
    def run(self):
        # try:
        while True:
            # print "avant ni_exe"
            # from subprocess import Popen
            # Popen(self.exe)


            from subprocess import check_output
            check_output(self.exe, shell=True)
            time.sleep(float(self.temp_attendre))
        


class Our_Tools(threading.Thread):

    def prestation_chps(
        self
        , code_prestation = "SGC"
        , client = "SOGEC"
        , nom_prestation = "AQ13"
    ):
        # testena we SGC sa SFL ny code_prestation
        # # si code_prestation == SGC > table_jerena = vivetic_prestation
        # # si code_prestation == SFL > table_jerena = sgc_prestation
        # # si code_prestation == autre >> mbola tsy voa_geree
        #
        # prestation_id = (client, code_prestation, nom_prestation)
        # sous_dossier_id = 
        #
        # si code_prestation = SGC > vivetic_champs... upto prestation_id
        # si code_prestation = SFL > sgc_champs... upto prestation_id

        type001 = ""
        if (code_prestation == 'SGC'):
            type001 = "vivetic"
            pass
        elif (code_prestation == 'SFL'):
            type001 = "sgc"            
            pass
        else:
            print "Not yet managed"
            sys.exit(0)
            pass

        table_jerena = type001 + "_prestation"
        req_presta_sous_doss = "SELECT "+type001+"_prestation_id AS prestation_id, sous_dossier_id AS sous_dossier_id FROM " + table_jerena + " WHERE nom_prestation = '" + nom_prestation + "' AND client = '" + client + "' AND code_prestation = '" + code_prestation + "' "
        # print "req_presta_sous_doss: ", req_presta_sous_doss
        # # req_presta_sous_doss:  SELECT vivetic_prestation_id as prestation_id, sous_dossier_id as sous_dossier_id from vivetic_pr estation where nom_prestation = 'AQ13' and client = 'SOGEC' and code_prestation = 'SGC'
        prest_id__sous_doss_id = self.pg_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , database01 = parser.get('pg_10_5_production', 'database')
            , query = req_presta_sous_doss
            , msg_if_error = ""
        )

        prestation_id = prest_id__sous_doss_id[0][0]
        sous_dossier_id = prest_id__sous_doss_id[0][1]

        print "prestation_id: "
        Our_Tools.print_green (txt = prestation_id)
        # # 1706
        print "sous_dossier_id: "
        Our_Tools.print_green (txt = sous_dossier_id)
        # # 324

        req_chps = "SELECT * FROM "+type001+"_champs WHERE " + type001 + "_prestation_id = "+ str(prestation_id)+"; "

        chps = self.pg_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , database01 = parser.get('pg_10_5_production', 'database')
            , query = req_chps
            , msg_if_error = ""
        )

        # print "chps6549876531231: ", chps
        # # [(73415, 'date_cachet_poste', 'Date cachet de

        Our_Tools.long_print(num = 5)
        for lg_chp in chps:
            print "Libellee: ", lg_chp[1]
            print "Description: ", lg_chp[2]
            print ""
        pass

    def doing_suppr_gpao_unique(
        self
        , suppr_total = 0
        , cmd001 = "cmd001"
        , all_lots = "all_lots01"
    ):
        delete_query_prod001 = "DELETE FROM pli_numerisation WHERE id_lot_numerisation IN "
        delete_query_prod001 += "(SELECT id_lot_numerisation FROM lot_numerisation WHERE "

        print "suppr_total0345676: ", suppr_total
        print "type(suppr_total): ", type(suppr_total)
        # raw_input()

        if (suppr_total == 0):
            delete_query_prod001 += "lot_scan IN (" + all_lots + ")  AND"
        delete_query_prod001 +=" idcommande_reception IN ('"+cmd001+"','0"+cmd001+"'));"
        
        print "delete_query_prod001654987987: ", delete_query_prod001
        # raw_input()

        delete_query_prod002 = "DELETE FROM pli_numerisation_anomalie WHERE id_lot_numerisation IN "
        delete_query_prod002 += "(SELECT id_lot_numerisation FROM lot_numerisation where "
        if suppr_total == 0:
            delete_query_prod002 += " lot_scan IN (" + all_lots + ")  AND"
        delete_query_prod002 += " idcommande_reception IN ('"+cmd001+"','0"+cmd001+"'));"




        delete_query_prod003 = "DELETE FROM image_numerisation WHERE id_lot_numerisation IN "
        delete_query_prod003 += "(SELECT id_lot_numerisation FROM lot_numerisation WHERE"
        if suppr_total == 0:
            delete_query_prod003 += " lot_scan IN (" + all_lots + ")  AND"
        delete_query_prod003 += " idcommande_reception IN ('"+cmd001+"','0"+cmd001+"'));"


        delete_query_prod004 = "DELETE FROM fichesuiveuse_numerisation WHERE id_lot_numerisation IN "
        delete_query_prod004 += "(SELECT id_lot_numerisation FROM lot_numerisation WHERE "
        if suppr_total == 0:
            delete_query_prod004 += " lot_scan IN (" + all_lots + ")  AND"
        delete_query_prod004 += " idcommande_reception IN ('"+cmd001+"','0"+cmd001+"'));"

        delete_query_prod005 = "DELETE FROM lot_numerisation WHERE "
        if suppr_total == 0:
            delete_query_prod005 += " lot_scan IN (" + all_lots + ")  AND"
        delete_query_prod005 += " idcommande_reception IN ('"+cmd001+"','0"+cmd001+"');"

        list_query_delete_prod = [
            delete_query_prod001, 
            delete_query_prod002, 
            delete_query_prod003,
            delete_query_prod004, 
            delete_query_prod005
        ]


        i = 0
        for query_prod in list_query_delete_prod:
            # print "delete_query_prod00"+ str(i) +": " + query_prod
            # self.logging_n_print( 
                # txt = query_prod + "\n", 
                # type_log = "info"
            # )

            
            self.pg_not_select(
                query01 = query_prod,
                host = "192.168.10.5",
                db = "production",
                log_query = True
                , auto_commit = True
                , test001 = False
            )

            i += 1
            Our_Tools.long_print()




        delete_query_sdsi001 = "DELETE FROM pousse WHERE idprep IN (SELECT idprep FROM fichier WHERE "
        if suppr_total == 0:
            delete_query_sdsi001 += "lot_client IN (" + all_lots + ") AND"
        delete_query_sdsi001 += " idcommande IN ('"+cmd001+"','0"+cmd001+"'));"
        
        delete_query_sdsi002 = "DELETE FROM fichierimage WHERE idprep IN (select idprep FROM fichier WHERE "
        if suppr_total == 0:
            delete_query_sdsi002 += "lot_client IN (" + all_lots + ") AND"
        delete_query_sdsi002 += " idcommande IN ('"+cmd001+"','0"+cmd001+"'));"
        # delete_query_sdsi003 = "DELETE FROM fichierimage_base64 WHERE idprep IN (SELECT idprep FROM fichier WHERE lot_client IN (" + all_lots + ") AND idcommande IN ('"+cmd001+"','0"+cmd001+"'));"
        delete_query_sdsi004 = "DELETE FROM preparation WHERE idprep IN (SELECT idprep FROM fichier WHERE "
        if suppr_total == 0:
            delete_query_sdsi004 += "lot_client IN (" + all_lots + ") AND"
        delete_query_sdsi004 += " idcommande IN ('"+cmd001+"','0"+cmd001+"'));"
        
        delete_query_sdsi005 = "DELETE FROM fichier WHERE "
        if suppr_total == 0:
            delete_query_sdsi005 += "lot_client IN (" + all_lots + ") AND"
        
        delete_query_sdsi005 += " idcommande IN ('"+cmd001+"','0"+cmd001+"');"

        list_query_delete_sdsi = [
            delete_query_sdsi001, 
            delete_query_sdsi002, 
            # delete_query_sdsi003,
            delete_query_sdsi004, 
            delete_query_sdsi005
        ]




        txt001 = """
##################################################################
# Dans bdd(sdsi) pour la commande("""+cmd001

        txt001 += " _ suppr_total" if (suppr_total == 1) else ""

        txt001 +=""")
##################################################################
# 
# 

         """
        # self.logging_n_print( 
            # txt = txt001 ,
            # type_log = "info")
        for query_sdsi in list_query_delete_sdsi:
            # print "delete_query_sdsi00"+ str(i) +": " + query_sdsi
            # self.logging_n_print( 
                # txt = query_sdsi + "\n", 
                # type_log = "info"
            # )

            #eto
            self.pg_not_select(
                query01 = query_sdsi
                , host = parser.get('pg_10_5_sdsi', 'ip_host')
                , db = parser.get('pg_10_5_sdsi', 'database')
                , log_query = True
                , auto_commit = True
                , test001 = False
            )

            # i += 1
            Our_Tools.long_print()



        long_void = "\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n"
        # self.logging_n_print( 
            # txt = long_void,
            # type_log = "info")

        print long_void
        print "Fin"
        # sys.exit(0) 
        pass

    def select_after_suppr_gpao_unique(
            self
            , suppr_total = 0
            , cmd001 = "cmd001"
            , all_lots = "all_lots01"
        ):

        sel_req_prod001 = "SELECT * FROM pli_numerisation WHERE id_lot_numerisation IN  (SELECT id_lot_numerisation FROM lot_numerisation WHERE "
        if suppr_total == 0:
            sel_req_prod001 += "lot_scan IN (" + all_lots + ") AND"
        sel_req_prod001 += " idcommande_reception IN ('" + cmd001 + "'));"


        sel_req_prod002 = "select * from pli_numerisation_anomalie where id_lot_numerisation in  (select id_lot_numerisation from lot_numerisation where "
        if suppr_total == 0:
            sel_req_prod002 += "lot_scan IN (" + all_lots + ") AND"
        sel_req_prod002 += " idcommande_reception IN ('" + cmd001 + "'));"


        sel_req_prod003 = "select * from image_numerisation where id_lot_numerisation in  (select id_lot_numerisation from lot_numerisation where "
        if suppr_total == 0:
            sel_req_prod003 += "lot_scan IN (" + all_lots + ") AND"
        sel_req_prod003 += " idcommande_reception IN ('" + cmd001 + "'));"



        sel_req_prod004 = "select * from fichesuiveuse_numerisation where id_lot_numerisation in  (select id_lot_numerisation from lot_numerisation where "
        if suppr_total == 0:
            sel_req_prod004 += "lot_scan IN (" + all_lots + ") AND"
        sel_req_prod004 += " idcommande_reception IN ('" + cmd001 + "'));"


        sel_req_prod005 = "select * from lot_numerisation where id_lot_numerisation in  (select id_lot_numerisation from lot_numerisation where "
        if suppr_total == 0:
            sel_req_prod005 += "lot_scan IN (" + all_lots + ") AND"
        sel_req_prod005 += " idcommande_reception IN ('" + cmd001 + "'));"

        list_sel_req_prod = [
            sel_req_prod001
            , sel_req_prod002
            , sel_req_prod003
            , sel_req_prod004
            , sel_req_prod005
        ]



        for req_select_prod in list_sel_req_prod:
            txt_to_add001 = "production@" + parser.get('pg_10_5_sdsi', 'ip_host') + ": \n" + req_select_prod
            Our_Tools.write_append_to_file(
                path_file = self.log_query_db,
                txt_to_add = txt_to_add001
            )

            pass





        sel_req_sdsi001 = "SELECT * FROM pousse WHERE idprep IN (SELECT idprep FROM fichier WHERE "
        if suppr_total == 0:
            sel_req_sdsi001 += "lot_client IN (" + all_lots + ") AND "
        sel_req_sdsi001 += "idcommande IN ('"+cmd001+"', '0"+cmd001+"'));"


        sel_req_sdsi002 = "select * from fichierimage where idprep in (select idprep from fichier where "
        if suppr_total == 0:
            sel_req_sdsi002 += "lot_client IN (" + all_lots + ") AND "
        sel_req_sdsi002 += "idcommande IN ('"+cmd001+"', '0"+cmd001+"'));"

        sel_req_sdsi003 = "select * from preparation where idprep in (select idprep from fichier where "
        if suppr_total == 0:
            sel_req_sdsi003 += "lot_client IN (" + all_lots + ") AND "
        sel_req_sdsi003 += "idcommande IN ('"+cmd001+"', '0"+cmd001+"'));"


        sel_req_sdsi004 = "select * from fichier where idprep in (select idprep from fichier where "
        if suppr_total == 0:
            sel_req_sdsi004 += "lot_client IN (" + all_lots + ") AND "
        sel_req_sdsi004 += "idcommande IN ('"+cmd001+"', '0"+cmd001+"'));"




        list_sel_req_sdsi = [
            sel_req_sdsi001
            , sel_req_sdsi002
            , sel_req_sdsi003
            , sel_req_sdsi004
        ]



        for req_select_sdsi in list_sel_req_sdsi:
            txt_to_add001 = "sdsi@" + parser.get('pg_10_5_sdsi', 'ip_host') + ": \n" + req_select_sdsi
            Our_Tools.write_append_to_file(
                path_file = self.log_query_db,
                txt_to_add = txt_to_add001
            )

            pass

        pass

    @staticmethod
    def export_query_to_xl(
            host = parser.get('pg_10_5_production', 'ip_host'), 
            username = parser.get('pg_10_5_production', 'username'),
            password = parser.get('pg_10_5_production', 'password'),
            database = parser.get('pg_10_5_production', 'database') 
    ):
        
        pass



    
    def update_lot_client_du_cmd(
        self
        , with_prompt = False     # We alefa avy ary am prompt ny valeur ilaina rht
    ):
        # 0 nom_program
        # 1 -T
        # 2 suppr_CRH001
        # 3 lot_client_new_val
        # 4 lot_client_to_repl
        # 5 commande
        
        if (with_prompt == True):
            req_sdsi = "UPDATE fichier SET lot_client = '" + sys.argv[3] + "' WHERE lot_client = '" + sys.argv[4] + "' AND idfichiercmd ILIKE '" + sys.argv[5] + "%'"
            # req_update
            print "req_sdsi: ", req_sdsi
            req_prod = "UPDATE lot_numerisation SET lot_scan = '"+ sys.argv[3] +"' WHERE lot_scan ILIKE '"+ sys.argv[4] +"' AND idcommande_reception = '"+ sys.argv[5] +"';"
            print "req_prod: ", req_prod
            pass
        else:
            workbook_write = xlsxwriter.Workbook('upd_lot_client.xlsx')
            sheet_modif_lot_client = workbook_write.add_worksheet('Modification Lot Client')
            header_format_red = workbook_write.add_format({'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': '#c80815',
                    'border': 1})
            sheet_modif_lot_client.write(0, 0, 
                'Modification Lot Client', header_format_red)
            header_format_blue = workbook_write.add_format({'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': '#4d8fac',
                    'border': 1})
            sheet_modif_lot_client.write(3, 0, 'Commande', header_format_blue)
            sheet_modif_lot_client.write(3, 1, 'Lot Aa Remplacer', header_format_blue)
            sheet_modif_lot_client.write(3, 2, 'Nouvelle Lot', header_format_blue)




            sheet_modif_lot_client.set_column('A:A', 20)
            sheet_modif_lot_client.set_column('B:B', 75)

            workbook_write.close()
            os.system('upd_lot_client.xlsx')

            workbook_read = xlrd.open_workbook('upd_lot_client.xlsx')
            sheet_read = workbook_read.sheet_by_index(0)
            try:

                # time.sleep(10)
                                            #  y  x

                cmd001 = sheet_read.cell_value(4, 0)
                lot_client_to_repl = sheet_read.cell_value(4, 1)
                lot_client_new_val = sheet_read.cell_value(4, 2)

                if isinstance(cmd001, float):
                    cmd001 = str(cmd001)[:-2]
                if isinstance(lot_client_to_repl, float):
                    lot_client_to_repl = str(lot_client_to_repl)[:-2]
                if isinstance(lot_client_new_val, float):
                    lot_client_new_val = str(lot_client_new_val)[:-2]
            except IndexError:
                print
                Our_Tools.print_green(txt = "Vous avez entree vide quelque part")
                self.logging_n_print( 
                    txt = "Commande vide dans l'Excel\n\n"
                    , type_log = "info"
                )
                sys.exit(0)
            # print "cmd001: ", cmd001
            # print "lot_client_to_repl: ", lot_client_to_repl
            # print "lot_client_new_val: ", lot_client_new_val
            list_req_sdsi_prod__upd_lot_client = Our_Tools.query_for_upd_lot_client(
                cmd = cmd001
                , lot_client_to_repl = lot_client_to_repl
                , lot_client_new_val = lot_client_new_val
            )

            i = 0
            for request001 in list_req_sdsi_prod__upd_lot_client:
                if i == 0:  # requete ao am sdsi
                    self.pg_not_select(
                        host = parser.get('pg_10_5_sdsi', 'ip_host')
                        , db = parser.get('pg_10_5_sdsi', 'database')
                        , query01 = request001
                        , log_query = True
                        , auto_commit = True
                    )
                    pass

                else: # requete ao am prod
                    self.pg_not_select(
                        host = parser.get('pg_10_5_production', 'ip_host')
                        , db = parser.get('pg_10_5_production', 'database')
                        , query01 = request001
                        , log_query = True
                        , auto_commit = True
                    )
                    pass
                i = i + 1
            



            list_sel_query_of_upd_lot_client = Our_Tools.select_query_for_upd_lot_client(
                        cmd = cmd001
                        , lot_client = lot_client_new_val
            )

            i = 0
            for sel_query_of_upd_lot_client in list_sel_query_of_upd_lot_client:
                if i == 0:  # SDSI
                    txt_to_add = "sdsi@" + parser.get('pg_10_5_sdsi', 'ip_host') + ": \n" + sel_query_of_upd_lot_client
                    pass
                else: 
                    txt_to_add = "production@" + parser.get('pg_10_5_sdsi', 'ip_host') + ": \n" + sel_query_of_upd_lot_client
                    pass

                self.write_append_to_file(
                    path_file = self.log_query_db
                    , txt_to_add = txt_to_add
                )

                i += 1

            self.write_append_to_file(
                path_file = self.log_query_db
                , txt_to_add = "\n\n\n\n\n"
            )




            
                
    

    @staticmethod
    def select_query_for_upd_lot_client(
        cmd = "cmd001",
        lot_client = "lot_client001"
    ):
        req_sel_sdsi = "SELECT * FROM fichier WHERE lot_client = '" + lot_client + "' AND idfichiercmd ILIKE '"+ cmd + "%'"

        req_sel_prod = "SELECT * FROM lot_numerisation WHERE lot_scan = '"+ lot_client +"' AND idcommande_reception = '"+ cmd +"';"
        return [req_sel_sdsi, req_sel_prod]
        pass


    @staticmethod
    def query_for_upd_lot_client(
        cmd = "cmd001"
        , lot_client_new_val = "lot_client_new_val"
        , lot_client_to_repl = "lot_client_to_repl"
    ):
        req_sdsi = "UPDATE fichier SET lot_client = '" + lot_client_new_val + "' WHERE lot_client = '" + lot_client_to_repl + "' AND idfichiercmd ILIKE '" + cmd + "%'"

        req_prod = "UPDATE lot_numerisation SET lot_scan = '"+ lot_client_new_val +"' WHERE lot_scan ILIKE '"+ lot_client_to_repl +"' AND idcommande_reception = '"+ cmd +"';"

        return [req_sdsi, req_prod]
        pass


    @staticmethod
    def refactor_sfl_correspondance():
        # this is Hasina_IAM_Program
        try:
            sous_doc = sys.argv[3]
        except IndexError:
            Our_Tools.long_print(num = 10)
            Our_Tools.print_red(txt = "Veuillez donnee l_idsousdossier dans le parametre")
            print "- ex:"
            Our_Tools.print_blue(txt = "> python Our_Tools.py -T test_selenium_sfl AP03")
            return
        rechercher = sous_doc
        chromedriverexe = parser.get('softw_path_exe', 'selenium_chromedriver')
        chromeexe = parser.get('softw_path_exe', 'chrome_exe')
        chromeOps = webdriver.ChromeOptions()
        chromeOps._binary_location = chromeexe
        print "02"                    
        driver = webdriver.Chrome(chromedriverexe, chrome_options=chromeOps)
        driver.get(
            parser.get('selenium_correspondance_sfl', 'link_bouygues')
        ) # Load page
        rch=driver.find_element_by_xpath("/html/body/form/table/tbody/tr[2]/td[2]/select")
        slct = rch.find_element_by_xpath("option[text()='" + rechercher + "']")
        slct.click()
        matr = driver.find_element_by_xpath("/html/body/form/table/tbody/tr[3]/td[2]/input")
        login = '99999'
        matr.send_keys(login)
        matr.send_keys(Keys.RETURN)                                        
        consulter = driver.find_element_by_xpath("/html/body/div[3]/table/tbody/tr[2]/td[2]/a")
        consulter.click()                    
        nvldmd = driver.find_element_by_xpath("//*[@id='nouvelleDemande']")
        nvldmd.click()                    
        liste_parent = driver.find_elements_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr")                                        
        t_name=[]
        i=0
        print len(liste_parent)
        for t in range(len(liste_parent)-1):
            i+=1
            try:                            
                d = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[2]/input")
                print d.get_attribute('name')
                d_str = str(d.get_attribute('name'))
                d1 = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[1]")
                d2_str = d1.text
                if (d_str.find('password')==-1):
                    t_name.append(d_str+"\t"+d2_str)                            
            except:
                try:
                    f = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[2]/label[1]/input")
                    print f.get_attribute('name')
                    f_str = str(f.get_attribute('name'))
                    f1 = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[1]")
                    f2_str = f1.text                                
                    if (f_str.find('password')==-1):
                        t_name.append(f_str+"\t"+f2_str)                                
                except:
                    g = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[2]/select")
                    print g.get_attribute('name')
                    g_str = str(g.get_attribute('name'))
                    g1 = driver.find_element_by_xpath("/html/body/div[3]/form/table[1]/tbody/tr["+str(i)+"]/td[1]")
                    g2_str = g1.text                                
                    t_name.append(g_str+"\t"+g2_str)                                        
        fichier = open("name_SFL_"+rechercher+".txt", "a")
        for t in t_name:
            fichier.write("%s\n"%(t))                    
        fichier.close()                    
        print 'DONE!'                    
        driver.quit()
        pass

    @staticmethod
    def test_tabwidget_pyqt001():
        class Test001(QTabWidget):
            def __init__(self, title, parent):
                # QtGui.QDockWidget.__init__(self, title, parent)
                QDockWidget.__init__(self, title, parent)
                self.parent = parent
                self.currentLocation = None

                # self.tabWidget = QtGui.QTabWidget()
                self.tabWidget = QTabWidget()
                # self.tabWidget.setTabPosition(QtGui.QTabWidget.West)
                self.tabWidget.setTabPosition(QTabWidget.West)
                self.setWidget(self.tabWidget)
                
                dropArea = DropArea(hostTypes.keys())
                dropArea2 = DropArea(netTypes.keys())
                dropArea3 = DropArea(customTypes.keys())
                self.tabWidget.addTab(dropArea, self.tr("&Host Element"))
                self.tabWidget.addTab(dropArea2, self.tr("&Net Element"))
                self.tabWidget.addTab(dropArea3, self.tr("&Custom Element"))
                
                self.connect(self,
                             QtCore.SIGNAL("dockLocationChanged(Qt::DockWidgetArea)"),
                             self.locationChanged)
                self.connect(self.tabWidget,
                             QtCore.SIGNAL("currentChanged(int)"),
                             self.tabChanged)
        app = QApplication(sys.argv)
        test001 = Test001("Title001", "Parent001")
        test001.show()
        sys.exit(app.exec_())

        pass

    @staticmethod
    def test_tabwidget_pyqt():

        # https://www.tutorialspoint.com/pyqt/pyqt_qtabwidget.htm
        class tabdemo(QTabWidget):
           def __init__(self, parent = None):
              super(tabdemo, self).__init__(parent)
              self.tab1 = QWidget()
              self.tab2 = QWidget()
              self.tab3 = QWidget()
                
              self.addTab(self.tab1,"Tab 1")
              self.addTab(self.tab2,"Tab 2")
              self.addTab(self.tab3,"Tab 3")
              self.tab1UI()
              self.tab2UI()
              self.tab3UI()
              self.setWindowTitle("tab demo")
                
           def tab1UI(self):
              layout = QFormLayout()
              layout.addRow("Name",QLineEdit())
              layout.addRow("Address",QLineEdit())
              self.setTabText(0,"Contact Details")
              self.tab1.setLayout(layout)
                
           def tab2UI(self):
              layout = QFormLayout()
              sex = QHBoxLayout()
              sex.addWidget(QRadioButton("Male"))
              sex.addWidget(QRadioButton("Female"))
              layout.addRow(QLabel("Sex"),sex)
              layout.addRow("Date of Birth",QLineEdit())
              self.setTabText(1,"Personal Details")
              self.tab2.setLayout(layout)
                
           def tab3UI(self):
              layout = QHBoxLayout()
              layout.addWidget(QLabel("subjects")) 
              layout.addWidget(QCheckBox("Physics"))
              layout.addWidget(QCheckBox("Maths"))
              self.setTabText(2,"Education Details")
              self.tab3.setLayout(layout)
                
        def main():
           app = QApplication(sys.argv)
           ex = tabdemo()
           ex.show()
           sys.exit(app.exec_())

        main()

    @staticmethod
    def test_multithread005():
        import logging
        import threading
        import time
        
        logging.basicConfig(level=logging.DEBUG,
                            format='(%(threadName)-10s) %(message)s',
                            )
                            
        def wait_for_event(e):
            """Wait for the event to be set before doing anything"""
            logging.debug('wait_for_event starting')
            event_is_set = e.wait()
            logging.debug('event set: %s', event_is_set)
        
        def wait_for_event_timeout(e, t):
            """Wait t seconds and then timeout"""
            while not e.isSet():
                logging.debug('wait_for_event_timeout starting')
                event_is_set = e.wait(t)
                logging.debug('event set: %s', event_is_set)
                if event_is_set:
                    logging.debug('processing event')
                else:
                    logging.debug('doing other work')
        
        
        e = threading.Event()
        t1 = threading.Thread(name='block', 
                            target=wait_for_event,
                            args=(e,))
        t1.start()
        
        t2 = threading.Thread(name='non-block', 
                            target=wait_for_event_timeout, 
                            args=(e, 2))
        t2.start()
        
        logging.debug('Waiting before calling Event.set()')
        time.sleep(3)
        e.set()
        logging.debug('Event is set')

    @staticmethod
    def test_multithread004():
        import threading
        import time
        import logging
        
        logging.basicConfig(level=logging.DEBUG,
                            format='(%(threadName)-10s) %(message)s',
                            )
        
        def daemon():
            logging.debug('Starting')
            time.sleep(2)
            logging.debug('Exiting')
        
        d = threading.Thread(name='daemon', target=daemon)
        d.setDaemon(True)
        
        def non_daemon():
            logging.debug('Starting')
            logging.debug('Exiting')
        
        t = threading.Thread(name='non-daemon', target=non_daemon)
        
        d.start()
        t.start()
        pass

    # https://pymotw.com/2/threading/
    @staticmethod
    def test_multithread003():
        import logging
        import threading
        import time

        logging.basicConfig(level=logging.DEBUG,
                format='[%(levelname)s] (%(threadName)-10s) %(message)s',
        )

        def worker():
            logging.debug('Starting')
            time.sleep(2)
            logging.debug('Exiting')

        def my_service():
            logging.debug('Starting')
            time.sleep(3)
            logging.debug('Exiting')

        t = threading.Thread(name='my_service', target=my_service)
        w = threading.Thread(name='worker', target=worker)
        w2 = threading.Thread(target=worker) # use default name

        w.start()
        w2.start()
        t.start()
        pass

    # https://pymotw.com/2/threading/
    @staticmethod
    def test_multithread002():
        import threading

        def worker():
            """thread worker function"""
            print 'Worker'
            return
        
        threads = []
        for i in range(5):
            t = threading.Thread(target=worker)
            threads.append(t)
            t.start()
            time.sleep(1)
        pass

    # https://www.tutorialspoint.com/python/python_multithreading.htm
    # https://geo.mydati.com/partners/HSS-6.9.1-install-hss-826-plain.exe
    @staticmethod
    def test_multithread001():
        import Queue
        import threading
        import time

        exitFlag = 0


        class myThread (threading.Thread):
           def __init__(self, threadID, name, q):
              threading.Thread.__init__(self)
              self.threadID = threadID
              self.name = name
              self.q = q
           def run(self):
              print "Starting " + self.name
              process_data(self.name, self.q)
              print "Exiting " + self.name

        def process_data(threadName, q):
           while not exitFlag:
                queueLock.acquire()
                if not workQueue.empty():
                    data = q.get()
                    queueLock.release()
                    print "%s processing %s" % (threadName, data)
                else:
                    queueLock.release()
                time.sleep(1)


        threadList = ["Thread-1", "Thread-2", "Thread-3"]
        nameList = ["One", "Two", "Three", "Four", "Five"]
        queueLock = threading.Lock()
        workQueue = Queue.Queue(10)
        threads = []
        threadID = 1

        # Create new threads
        for tName in threadList:
           thread = myThread(threadID, tName, workQueue)
           thread.start()
           threads.append(thread)
           threadID += 1

        # Fill the queue
        queueLock.acquire()
        for word in nameList:
           workQueue.put(word)
        queueLock.release()

        # Wait for queue to empty
        while not workQueue.empty():
           pass

        # Notify threads it's time to exit
        exitFlag = 1

        # Wait for all threads to complete
        for t in threads:
           t.join()
        print "Exiting Main Thread"
        pass

    def reformat_thread_conf_test_connect_prod_10_5(
        self, 
        host = parser.get(
            'pg_10_5_production', 
            'ip_host'),
        database = parser.get(
            'pg_10_5_production', 
            'database')
    ):
        if self.db_is_connected(
            host = host,
            database01 = database
        ):
            # print "connection ok au bdd(production)"
            Our_Tools.write_append_to_file(
                path_file = path_prg + "log_connect_db.txt",
                txt_to_add = str(datetime.datetime.now()) + ": Connection OK au bdd("+database+")@"+host
            )
            pass
        else:
            
            txt001 = "************ " + str(datetime.datetime.now()) + ": Connection PERDU pour bdd("+database+")@" + host
            txt002 = str(datetime.datetime.now()) + ": Connection PERDU pour bdd("+database+")@" + host
            Our_Tools.write_append_to_file(
                path_file = path_prg + "log_connect_db.txt",
                txt_to_add = txt001
            )
            Our_Tools.popup(
                window_title = "Erreur de connection de base",
                msg = txt002)
        pass

    @staticmethod
    def copy_dir_content(
            path_src = 'E:\DISK_D\date',
            path_target = 'E:\DISK_D\ASA\BLF') :
        from distutils.dir_util import copy_tree

        # copy subdirectory example
        # fromDirectory = "/a/b/c"
        # toDirectory = "/x/y/z"

        copy_tree(path_src, path_target)

    @staticmethod
    def long_print(num = 10):
        for i in range(num):
            print
        pass

    @staticmethod
    def edit_file_w_sublime2(
            path_file = "E:\\DISK_D\\mamitiana\\kandra\\do_not_erase\\our_tools\\" + "redmine_file.txt"):
        # subprocess.Popen(["subl", "-w", path_file]).wait()
        # path_sublime2 = "C:\Program Files\Sublime Text 2\sublime_text.exe"
        subprocess.Popen([path_sublime2, "-w", path_file]).wait()
        print path_file
        pass

    @staticmethod
    def read_file_line_by_line(path_file = "ti.txt"):
        res = ""
        with open(path_file) as open_file_path:
            for line in open_file_path:
                res += line
        return res

    def db_is_connected(self,
            host = "192.168.10.5",
            database01 = 'production'):

        res = False

        self.pg_select(
            host = host,
            database01 = database01,
            query = "SELECT 1"
        )
        if (self.rows_pg_10_5__prod[0][0] == 1):
            # time.sleep(self.time_test_connection)
            # print "connection ok"
            res = True

        return res

        pass



    @staticmethod
    def replacer_factory(spelling_dict):
        def replacer(match):
            word = match.group()
            return spelling_dict.get(word, word)
        return replacer

    @staticmethod
    def replace_in_file(
            path_file_input = "ti.txt",
            path_file_output = "ta.txt",
            replacements = {
                "coco": 'toto',
                "tic": """tac
                this is going to be a very long txt
                hereis not yet the end
                looks like this is the end
                finally, the end"""
            }):
        with open(path_file_input) as infile, open(path_file_output, 'w') as outfile:
            for line in infile:
                for src, target in replacements.iteritems():
                    line = line.replace(src, target)
                outfile.write(line)
        pass

    @staticmethod
    def manage_usb_store_w_regedit(state = 3):
        # state = 3 > activated
        # state = 4 > deactivated

        # this has to be run as an administrator_account
        keyVal = r'SYSTEM\CurrentControlSet\services\USBSTOR'

        try:
            key = OpenKey(HKEY_LOCAL_MACHINE, 'SYSTEM\\CurrentControlSet\\services\\USBSTOR', 
            0
            , KEY_ALL_ACCESS      # this is going to create error if you ran the method without admin_account
            )

            # SetValueEx(key, "Start Page", 0, REG_SZ, "http://www.google.com/")
            SetValueEx(key, "Start", 0, REG_DWORD, state)

            if state == 3:
                Our_Tools.print_green(
                    txt = 'USB_storage Activated',
                    new_line = False)
                pass

                Our_Tools.popup(
                    window_title = "USBSTOR",
                    msg = "Veuillez ReConnecter votre USB pour l_Activee")

            elif state == 4:
                # Our_Tools.print_green(
                    # txt = 'USB_storage DeActivated',
                    # new_line = False)

                Our_Tools.print_green(
                    txt = 'USB_storage',
                    new_line = False)
                Our_Tools.print_red(
                    txt = 'De',
                    new_line = False)
                Our_Tools.print_green(
                    txt = 'Activated',
                    new_line = False)

                Our_Tools.popup(
                    window_title = "USBSTOR",
                    msg = "Veuillez ReConnecter votre USB pour le DesActivee")

                pass

            CloseKey(key)

            pass
            # SetValueEx(key, "Start", 0, REG_DWORD, str(state))
            # CloseKey(key)
        except WindowsError:
            Our_Tools.long_print()
            Our_Tools.print_red(
                txt = 'Vous devez executer cette commande en compte_Admin 654987312344566',
                new_line = False)
            # print "this is a test"
            pass
        except Exception:
            Our_Tools.long_print()
            Our_Tools.print_red ("there is error 9876568965654654")
            sys.exit(0)
            key = CreateKey(HKEY_CURRENT_USER, keyVal)
        
        # print "done 6549316876"

        pass

    @staticmethod
    def ame_to_bre(text = "this is a test",
        changmt = {
            'tire':'tyre', 
            'color':'colour', 
            'utilize':'utilise',
            'foo': 'bar'
            }):
        pattern = r'\b\w+\b'  # this pattern matches whole words only

        replacer = Our_Tools.replacer_factory(changmt)
        return re.sub(pattern, replacer, text)

    def connection_pg(self,  
            server01 = parser.get('pg_localhost_saisie', 'ip_host'),
            user01=parser.get('pg_localhost_saisie', 'username'),
            password01=parser.get('pg_localhost_saisie', 'password'),
            database01=parser.get('pg_localhost_saisie', 'database')
    ):
        try:
            # host = 10.5
            if(server01 == parser.get('pg_10_5_sdsi', 'ip_host')):
                if (database01 == parser.get('pg_10_5_sdsi', 'database')):
                    try:
                        self.connect_pg_10_5_sdsi
                    except AttributeError:
                        self.connect_pg_10_5_sdsi = psycopg2.connect(
                            "dbname=" + database01
                            +" user=" + user01
                            +" password=" + password01
                            +" host=" + server01
                        )
                        self.connect_pg_10_5_sdsi.set_isolation_level(0)
                        self.cursor_pg_10_5__bdd_sdsi = self.connect_pg_10_5_sdsi.cursor()
                        print "Connection OK au pg10.5 bdd(sdsi)"


                # host = 10.5 
                elif (database01 == parser.get('pg_10_5_production', 'database')):
                    self.connect_pg_10_5__prod = psycopg2.connect(
                        "dbname=" + database01
                        +" user=" + user01
                        +" password=" + password01
                        +" host=" + server01
                    )
                    self.connect_pg_10_5__prod.set_isolation_level(0)
                    self.cursor_pg_10_5__bdd_prod = self.connect_pg_10_5__prod.cursor()
                    print "Connection OK au pg "+server01+" bdd("+database01+")"

            elif(
                    (server01 == parser.get('pg_localhost_saisie', 'ip_host'))
                    and (database01 == parser.get('pg_localhost_saisie', 'database'))
            ):
                try:
                    self.connect_pg_localhost_saisie
                except AttributeError:
                    self.connect_pg_localhost_saisie = psycopg2.connect(
                        "dbname=" + database01
                        +" user=" + user01
                        +" password=" + password01
                        +" host=" + server01
                    )
                    self.connect_pg_localhost_saisie.set_isolation_level(0)
                    self.cursor_pg_localhost_saisie = self.connect_pg_localhost_saisie.cursor()
                    print "Connection ok au bdd(saisie)@localhost"
                
                pass
            # host = 127.0.0.1
            elif(server01 == parser.get('pg_localhost_sdsi', 'ip_host')):
                if (database01 == parser.get('pg_localhost_sdsi', 'database')):
                    try:
                        self.connect_pg_local_sdsi
                    except AttributeError:
                        self.connect_pg_local_sdsi = psycopg2.connect(
                            "dbname=" + database01
                            +" user=" + user01
                            +" password=" + password01
                            +" host=" + server01
                        )
                        self.connect_pg_local_sdsi.set_isolation_level(0)
                        self.cursor_pg_local__bdd_sdsi = self.connect_pg_local_sdsi.cursor()
                        print "Connection OK au pg_localhost bdd(sdsi_local)"
                pass
            elif(
                    (server01 == parser.get('pg_10_32_production', 'ip_host'))
                    and (database01 == parser.get('pg_10_32_production', 'database'))
            ):
                try:
                    self.connect_pg_10_32_prod
                except AttributeError:
                    self.connect_pg_10_32_prod = psycopg2.connect(
                        "dbname=" + database01
                        +" user=" + user01
                        +" password=" + password01
                        +" host=" + server01
                    )
                    self.connect_pg_10_32_prod.set_isolation_level(0)
                    self.cursor_pg_10_32_prod = self.connect_pg_10_32_prod.cursor()
                    print "Connection OK au pg_10_32 bdd(production)"
                

                pass
        except(psycopg2.OperationalError):
            print ""
            print ""
            print ""
            print "there is an OperationalError"
            # txt = 
            Our_Tools.write_append_to_file(
                    path_file = 'log_connect_db.txt',
                    txt_to_add = str(datetime.datetime.now()) + ": Impossible de se connecter au db__" + database01)
            Our_Tools.popup(
                window_title = "Connection interrompue",
                msg = "Connection interrompue sur la base de donnee(" + database01 + ")")
            # self.connection_pg(
                # server01 = server01,
                # user01 = user01,
                # password01 = password01,
                # database01 = database01)
            time.sleep(1)

    def logging_n_print(self, 
            bool01 = True,
            type_log = "info", # OU warning OU debug
            txt = "text",
            log_only = True):

        if(type_log == "warning"):
            logging.warning(txt)
            if log_only == False:
                print txt
            

        elif (type_log == "info"):
            logging.info(txt)
            if log_only == False:
                print txt

        elif ((type_log == "debug") & (log_only == False)):
            logging.debug(txt)
            if log_only == False:
                print txt

        elif ((type_log == "debug") & (log_only == True)):
            logging.debug(txt)
            if log_only == False:
                print txt


    @staticmethod
    def get_xls_tab_name(
        path_file_xls = 'AG22\\AG08_DescripteurSaisie.xls'
        ):
        pointSheetObj = []
        xl_read = xlrd.open_workbook(path_file_xls)
        pointSheets = xl_read.sheet_names()

        # print pointSheets
        return pointSheets

        # otrn tsy ilaina ireo ambany ireo aah
        # sys.exit(0)

        # for i in pointSheets:
            # pointSheetObj.append(
                # tuple(
                    # (xl_read.sheet_by_name(i),i)))
# 
        # res = []
# 
        # for obj, sh_n in pointSheetObj:
            # res.append(sh_n)
# 
        # return res

    @staticmethod
    def test001():
        try:
            while True:
                print "test"
                time.sleep(1)
        except KeyboardInterrupt:
            print "interrompue"
        pass


    def suppression_gpao_unique(
            self
            , save_check_query = False
            , suppr_total = 0
    ):
        print "suppression_gpao_unique"
        # print "parser.get('pg_10_5_production', 'ip_host')"
        # print parser.get('pg_10_5_production', 'ip_host')
        # sys.exit(0)

        #ato
        # self.connection_pg(
            # server01 = parser.get('pg_localhost_sdsi', 'ip_host'),
            # user01=parser.get('pg_localhost_sdsi', 'username'),
            # password01=parser.get('pg_localhost_sdsi', 'password'),
            # database01=parser.get('pg_localhost_sdsi', 'database')
        # )
# 
# 
        # sys.exit(0)

        self.connection_pg(
            server01 = parser.get('pg_10_5_production', 'ip_host'),
            user01=parser.get('pg_10_5_production', 'username'),
            password01=parser.get('pg_10_5_production', 'password'),
            database01=parser.get('pg_10_5_production', 'database')
            )
        self.connection_pg(
            server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
            user01=parser.get('pg_10_5_sdsi', 'username'),
            password01=parser.get('pg_10_5_sdsi', 'password'),
            database01=parser.get('pg_10_5_sdsi', 'database')
            )

        self.fichier_xlsx = "file01.xlsx"


        # # workbook_write = xlwt.Workbook()
        # if (os.path.exists(self.fichier_xlsx)):
        workbook_write = xlsxwriter.Workbook(self.fichier_xlsx)
        # workbook_write = xlwt.open_workbook_write(self.fichier_xlsx)
        # # sheet_write = workbook_write.add_sheet('suppression_gpao_unique')
        
        header_format_red = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#c80815',
                            'border': 1})


        header_format_blue = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#4d8fac',
                            'border': 1})

        sheet_suppr_gpao_uniq = workbook_write.add_worksheet('Suppression GPAO Unique')
        cell_format_union = workbook_write.add_format({'align': 'center',
            'valign': 'vcenter',
            'border': 1})
        sheet_suppr_gpao_uniq.merge_range('A1:B1', "", 
            cell_format_union)
        #                           y  x
        sheet_suppr_gpao_uniq.write(0, 0, 'Veuillez entrer les valeurs correspondants au dessous des cases colorees', header_format_red)
        sheet_suppr_gpao_uniq.write(3, 0, 'Commande', header_format_blue)
        sheet_suppr_gpao_uniq.write(3, 1, 'Lots', header_format_blue)
        
        
        
        sheet_suppr_gpao_uniq.set_column('A:A', 20)
        sheet_suppr_gpao_uniq.set_column('B:B', 75)


        # sheet_write.workbook_write.sheet_by_index(0)

        #                 y  x  " eto le txt soratana"
        # sheet_write.write(0, 0, "Ce ligne contient l_Utilisation de ce programme")
        # sheet_write.write(1, 0, "")

        # workbook_write.save(self.fichier_xlsx)
        workbook_write.close()
        os.system(self.fichier_xlsx)
        
        # vita soratra ny "commande" sy ny "lot" rhf tonga eto

        workbook_read = xlrd.open_workbook(self.fichier_xlsx)
        sheet_read = workbook_read.sheet_by_index(0)
        try:

        # time.sleep(10)
                                        #  y  x
            cmd001 = sheet_read.cell_value(4, 0)
        except IndexError:
            print
            Our_Tools.print_green("Vous avez entree vide pour la cellule(Commande)")
            self.logging_n_print( 
                txt = "Commande vide dans l'Excel\n\n",
                type_log = "info"
            )
            sys.exit(0)

        # print "cmd001: " + cmd001
        # # LOX098
        # sys.exit(0)

        # datas = [sheet_read.cell_value(1, col) for col in range(sheet_read.ncols)]
        #                               y   x                   
        datas = [sheet_read.cell_value(row, 1) for row in range(4, sheet_read.nrows)]
        # print datas
        # # [u'BE_20170923_02_P_R', u'BE_20170923_03_P_R', u'BE_20, ....
        
        # print "len(datas): " + str(len(datas))
        # print datas

        if (len(datas) == 1) and (len(datas[0]) == 0):  # possible we vide ny lot apdirn fa 
            print "tsisy nininn ao am lot_s aah"
            sys.exit(0)

        all_lots = ""
        for lot01 in datas:
            if isinstance(lot01, float):
                all_lots += "'" + '{:.0f}'.format(lot01) + "', "
            elif (isinstance(lot01, unicode) or isinstance(lot01, str) ):
                all_lots += "'" + str(lot01) + "', "

            # # 'BE_20170923_02_P_R','BE_20170923_03_P_R', ...
            # # no tanjona


        # all_lots = "','".join(datas)
        # all_lots = "'" + all_lots + "'"
        all_lots = all_lots[:-2]

        print all_lots
        # sys.exit(0)
        # print all_lots
        # # 'BE_20170923_02_P_R','BE_20170923_03_P_R', ...

       
        txt001 = """
# ################################################################
# Dans bdd(production) pour la commande("""+cmd001

        txt001 += " _ suppr_total" if (suppr_total == 1) else ""
        txt001 += """)
# ################################################################
# 
# 
         """
        self.logging_n_print( 
            txt = txt001,
            type_log = "info"
        )


        Our_Tools.long_print()
        # lot_scan IN (" + all_lots + ")  AND





        #######################################################
        #######################################################
        #######################################################
        #######################################################
        #######################################################
        #######################################################
        
        self.doing_suppr_gpao_unique(
            suppr_total = suppr_total
            , cmd001 = cmd001
            , all_lots = all_lots
        )

        if save_check_query:
            self.select_after_suppr_gpao_unique(
                suppr_total = suppr_total
                , cmd001 = cmd001
                , all_lots = all_lots
            )
        #######################################################
        #######################################################
        #######################################################
        #######################################################
        #######################################################
        #######################################################
        #######################################################



        
    @staticmethod
    def read_one_cell_from_xl(
        xl_file = "test001.xlsx"
        , sheet_index = 0
        , y = 2
        , x = 0
        , give_default_value_if_void = 1
    ):
        workbook_read = xlrd.open_workbook(xl_file)
        sheet_read = workbook_read.sheet_by_index(sheet_index)
        res = ""

        try:
            res = sheet_read.cell_value(y, x)
            if isinstance(res, float):
                res = '{:.0f}'.format(res)
            if ((give_default_value_if_void == 1) and (len(res) == 0)):
                res = 'Erreur dans "xl_file": ' + str(xl_file)+', "sheet_index": ' + str(sheet_index) + ", x = " + str(x) + ", y = " + str(y)
                res += '\n- Contenu du cellule vide'
            return res
        except IndexError:
            msg = "Valeur manquant pour fichier(" + xl_file + "), tab_index("+ sheet_index +")" + ", y = " + y + ", x = " + x
            Our_Tools.print_green(
                txt = msg
            )
            return False
            pass

        pass


    def refactor_sgc_operation_codebarre(
            self
            , has_code_barre_operation = 'has_code_barre_operation'
            , nom_prestation = 'nom_prestation'
    ):

        if has_code_barre_operation == '1':
            # print "Traitement code barre"
            # max_code_barre_id_plus_1
            req_max_code_barre_id_et_sgc_chp_inter_plus_1 = 'SELECT max(sgc_codebarre.sgc_codebarre_id)+1 codebarre,  max(sgc_champs_interdep.sgc_champs_interdep_id)+1 interdep FROM sgc_champs_interdep, sgc_codebarre'
            list_max_code_barre_id_et_sgc_chp_inter_plus_1 = self.pg_select(
                host = parser.get('pg_10_5_production', 'ip_host'),
                database01 = parser.get('pg_10_5_production', 'database'),
                query = req_max_code_barre_id_et_sgc_chp_inter_plus_1
            )
            max_code_barre_id_plus_1 = list_max_code_barre_id_et_sgc_chp_inter_plus_1[0][0]
            max_sgc_chp_inter_plus_1 = list_max_code_barre_id_et_sgc_chp_inter_plus_1[0][1]
            # print "type(max_code_barre_id_plus_1): ", type(max_code_barre_id_plus_1)
            # print "max_sgc_chp_inter_plus_1: ", max_sgc_chp_inter_plus_1



            list_numero_code_barre = Our_Tools.read_one_col_of_sheet_xl(
                xl_file = self.sgc_xlsx
                , sheet_index_to_read = 3
                , x = 0
            )

            list_nom_code_barre = Our_Tools.read_one_col_of_sheet_xl(
                xl_file = self.sgc_xlsx
                , sheet_index_to_read = 3
                , x = 1
            )

            # print "list_numero_code_barre: ", list_numero_code_barre
            # # [u'azer', u'qsdf', u'wxcv']
            # print "list_nom_code_barre: ", list_nom_code_barre

            # max_code_barre_id_plus_1
            req_code_barre = "insert into sgc_codebarre values '"

            for i in list_numero_code_barre:
                req_code_barre += "('"+str (max_code_barre_id_plus_1) + "', '" + i + "', '"+ nom_prestation + "'),\n"
                max_code_barre_id_plus_1 += 1
                pass
            req_code_barre = req_code_barre[:-2]
            # print "req_code_barre: ", req_code_barre


            self.pg_not_select(
                host = parser.get('pg_10_5_production', 'ip_host')
                , db = parser.get('pg_10_5_production', 'database')
                , query01 = req_code_barre
                , log_query = True
                , auto_commit = False
                , test001 = True
            )

        elif has_code_barre_operation == '0':
            print "CodeBarre Null"
        else:
            print "Unknown val for code barre"

        pass

    def pg_not_select(self
            , query01 = ""
            , host = "127.0.0.1"
            , db = "sdsi"
            , log_query = False
            , auto_commit = False
            , test001 = True
    ):
        if( 
                (host == parser.get('pg_10_5_sdsi', 'ip_host')) 
                and (db == parser.get('pg_10_5_sdsi', 'database'))
        ):
            # try:
            try:
                self.connect_pg_10_5_sdsi
            except AttributeError:
                self.connect_pg_10_5_sdsi = psycopg2.connect(
                    "dbname=" + parser.get('pg_10_5_sdsi', 'database')
                    +" user=" + parser.get('pg_10_5_sdsi', 'username')
                    +" password=" + parser.get('pg_10_5_sdsi', 'password')
                    +" host=" + parser.get('pg_10_5_sdsi', 'ip_host')
                )
                self.connect_pg_10_5_sdsi.set_isolation_level(0)
                self.cursor_pg_10_5__bdd_sdsi = self.connect_pg_10_5_sdsi.cursor()
                print "Connection OK au pg10.5 bdd(sdsi)"

            if test001 == False:
                self.cursor_pg_10_5__bdd_sdsi.execute(query01.encode('ISO-8859-1'))

            if (auto_commit == True):
                print "Commited SDSI"
                self.connect_pg_10_5_sdsi.commit()
                pass
            else:
                Our_Tools.print_green(txt = "Pas encore Commitee")
                print "- ", query01
                pass

        elif ( 
                (host == parser.get('pg_localhost_saisie', 'ip_host')) 
                and (db == parser.get('pg_localhost_saisie', 'database')) 
        ):
            try:
                self.connect_pg_localhost_saisie
            except AttributeError:
                self.connect_pg_localhost_saisie = psycopg2.connect(
                    "dbname=" + parser.get('pg_localhost_saisie', 'database')
                    +" user=" + parser.get('pg_localhost_saisie', 'username')
                    +" password=" + parser.get('pg_localhost_saisie', 'password')
                    +" host=" + parser.get('pg_localhost_saisie', 'ip_host')
                )
                self.connect_pg_localhost_saisie.set_isolation_level(0)
                self.cursor_pg_localhost_saisie = self.connect_pg_localhost_saisie.cursor()
                print "Connection ok au bdd(saisie)@localhost"

            if test001 == False:
                self.cursor_pg_localhost_saisie.execute(query01.encode('ISO-8859-1'))
            

            pass
            if (auto_commit == True):
                self.connect_pg_localhost_saisie.commit()
                pass
            else:
                Our_Tools.print_green(txt = "Pas encore Commitee")
                print "- ", query01
                pass

        # elif ( 
                # (host == parser.get('pg_10_5_sdsi', 'ip_host')) 
                # and (db == parser.get('pg_10_5_sdsi', 'database')) 
        # ):
            # try:
                # self.connect_pg_10_5_sdsi
            # except AttributeError:  
                # self.connection_pg(    # on fait une connection aa la base car elle est inexistant
                    # # cette methode va definir self.cursor_pg_10_5__bdd_sdsi
                    # server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
                    # user01=parser.get('pg_10_5_sdsi', 'username'),
                    # password01=parser.get('pg_10_5_sdsi', 'password'),
                    # database01=parser.get('pg_10_5_sdsi', 'database')
                # )
            # self.cursor_pg_10_5__bdd_sdsi.execute(query01)
# 
            # if (auto_commit == True):
                # self.connect_pg_10_5__prod.commit()
                # pass
            # else:
                # Our_Tools.print_green(txt = "Pas encore Commitee")
                # print "- ", query01
                # pass

        elif ( 
                (host == parser.get('pg_10_5_production', 'ip_host'))
                and (db == parser.get('pg_10_5_production', 'database')) 
        ):
            try: # de meme que sdsi@10.5
                self.connect_pg_10_5__prod
            except AttributeError:
                self.connection_pg( # on fait une connection aa la base car elle est inexistant
                    server01 = parser.get('pg_10_5_production', 'ip_host'),
                    user01=parser.get('pg_10_5_production', 'username'),
                    password01=parser.get('pg_10_5_production', 'password'),
                    database01=parser.get('pg_10_5_production', 'database')
                )

            # print "Execute2334579944678906453"
            if test001 == False:
                self.cursor_pg_10_5__bdd_prod.execute(query01.encode('ISO-8859-1'))
                print 'test == False'


            if (auto_commit == True):
                print "Commit12347689997654"
                self.connect_pg_10_5__prod.commit()
                pass
            else:
                Our_Tools.print_green(txt = "Pas encore Commitee")
                print "- ", query01
                pass




        txt_to_add001 = ">>>>>>>>>>>> " + db + "@" + host + ": " + str(datetime.datetime.now())+ ": \n"+query01
        if log_query:
            Our_Tools.write_append_to_file(
                path_file = self.log_query_db,
                txt_to_add = txt_to_add001
            )
            pass

    def commit_all(self):
        
        self.commit_specific(connection = "pg_10_5_production")
        self.commit_specific(connection = "pg_10_5_sdsi")
        self.commit_specific(connection = "pg_localhost_saisie")

        pass

    def commit_specific(
        self
        , connection = "pg_10_5_production"
    ):
        if connection == "pg_10_5_production":
            try: # de meme que sdsi@10.5
                self.connect_pg_10_5__prod
                self.connect_pg_10_5__prod.commit()
            except:
                Our_Tools.print_red(
                    txt = "Connection au 10.5_production pas faite mais on a fait un Commit "
                )
                pass
            pass
        elif connection == "pg_10_5_sdsi":
            try: # de meme que sdsi@10.5
                self.connect_pg_10_5_sdsi
                self.connect_pg_10_5_sdsi.commit()
            except:
                Our_Tools.print_red(
                    txt = "Connection au 10.5_sdsi pas faite mais on a fait un Commit "
                )
                pass
            pass
        elif connection == "pg_localhost_saisie":
            try: # de meme que sdsi@10.5
                self.connect_pg_localhost_saisie
                self.connect_pg_localhost_saisie.commit()
            except:
                Our_Tools.print_red(
                    txt = "Connection au localhost_saisie pas faite mais on a fait un Commit "
                )
                pass
            pass
        pass

    @staticmethod
    def usage():
        if os.name == 'nt':
            os.system('cls')
        elif os.name == 'posix':
            os.system('clear')

        Our_Tools.long_print()

        print "Usage: "
        Our_Tools.print_green (txt = "Option: -h, --help")
        print "> Our_Tools.py -h"
        print "> Our_Tools.py --help"
        print "- - ces 2scripts font la meme chose"
        print "- - ils vont afficher cette Aide"



        Our_Tools.long_print(num = 5)


        Our_Tools.print_green (txt = "Option: -L, --long-print")
        print "> Our_Tools.py -L"
        print "- - pour afficher 70lignes vides qui vont immiter 'cls' ou 'clear'"
        print "- - "
        print "- - pour effacer l_ecran"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -c, --crawl")
        print 'il est preferable d_utiliser crawl01.sh avec les 5params'
        print 'dans cette programme, il va faire rien du tout'
        print '-'
        print '-'
        print "> Our_Tools.py -c /path/folder001/ pattern_search001"
        print "OU"
        print "> Our_Tools.py --crawl /path/folder001/ pattern_search001"
        print '- - cela va chercher "pattern_search001" dans le chemin(/path/folder001/)'

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -d, --directory")
        print "> Our_Tools.py -d"
        print "- - pour chercher un repertoire dans un ordi"
        print "- - "
        print "- - C_est la meme que:"
        print "> find -maxdepth 1 -iname '*nameXX*'-type d"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -s, --suppression_gpao_unique")
        print "> Our_Tools.py -s"
        print "- - pour la suppression de gpao unique"
        print "- - "
        print "- - "
        print "- - "
        print "- - Il est possible que la Suppression ne possede pas de "
        print '- - - lots_clients... ie, on a k1 seule commande et on suppr tout'
        Our_Tools.print_yellow(txt = "> Our_Tools.py -s --total")
        
        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T, --all_test")
        print "Pour faire des test que j_ai trouvee sur Internet"
        print
        print "Our_Tools.py -T dl link01 file_save"
        print "Soyez en sure que link01 est de type(http://XXX/YYY.qsd"


        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -S, --sgc")
        print "- Ceci est en cours de Developpement"
        print "- Finit mais pas encore testee"
        print "- Pour faire les traitements des Tickets SOGEC-SGC"
        print "- ex: python Our_Tools.py -S"
        print "- ex: python Our_Tools.py --sgc"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -e, --export_table")
        print "- Pour exporter le contenu de certain table dans du fichier_excel"
        print "- host, db, table, output"
        print "- ex: python Our_Tools.py -e 192.168.10.5 production RED001_S1 out.xlsx"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T manage_usb_storage activate")
        print '###########Ceci doit etre executee en tant que compte_Admin###########'
        print "- To Activate the USB_storage"
        print "Option: -T manage_usb_storage deactivate"
        print "- To DeActivate the USB_storage"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_thread_conn")
        print "Va tester la connection Infiniment"
        print "-"
        print "Pour pouvoir l_arreter, il faut faire"
        print "> taskkill /f /pid pid_001 "
        print "-"
        print "-"
        print "Pour avoir le pid_001... Souvent le thread_connection est la premiere"
        print '> tasklist | find /i "python"'
        print "-"
        print "-"
        print "-"
        print "- Autrement, si la seule programme qui utilise python "
        print "- - est celui-ci,... Pour arreter le programme, on fait"
        print '> taskkill /f /im "python.exe"'

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_thread_redmine")
        print "Va afficher un pop_up quand il est 15:00"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_edit_file_redmine")
        print "Va Editer le fichier qui va s_afficher au moment du pop_up__redmine"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_thread_conf")
        print "Le thread qui sont Configurees dans fichier_conf__all_confs.ini "
        print "- vont etre executees"
        #fin_usage

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_selenium_sfl AP03")
        print "Pour faire le Correspondance de la commande AP03"
        print "- le resultat sera dans un fichier comme: 'name_SFL_AP03.txt'"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T test_conn_db025")
        print "Pour faire du keylogging dans ton pc"
        print "- le fichier log sera log_stuffs.txt"

        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T export_one_query_select_to_excel")
        print "Pour Exporter une SEULE requete dans excel"
        print 'Ex: python .\Our_Tools.py -T export_one_query_select_to_excel "select * from mpo002_s1 limit 6" "test265.xlsx"'


        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T update_lot_client_du_cmd")
        print "Pour faire un m_a_j d_un lot_client"
        print "- Un fichier_excel va s_ouvrir et tout va etre dedans"


        Our_Tools.long_print(num = 5)

        Our_Tools.print_green (txt = "Option: -T prestation_chps client001 code_prestation001 nom_prestation001")
        print "Pour afficher le(prestation_id, sous_dossier_id)... et tout les champs du (prestation_id, sous_dossier_id)"
        print "-"
        print "- ex: -T SOGEC SGC AQ13"
        

        # end_usage

    @staticmethod
    def display_pic(
            path_pic = path_prg + "icon.png"
    ):
        # print path_pic
        # img = Image.open(path_pic)
        # img = Image.open(r'E:\DISK_D\mamitiana\kandra\do_not_erase\our_tools\icon.png')
        # # https://stackoverflow.com/questions/1413540/showing-an-image-from-console-in-python
        # # # not working in windows7
        # img = Image.open('icon.png')
        # img.show()

        if os.name == 'nt':
            os.system(path_pic)
            # os.system('cls')

        pass

    def manage_redmine_popup(self):
        # print "ty ooh"
        now_date_time = datetime.datetime.now()

        self.time_redmine_popup = "15:00:00"
        self.time_redmine_popup = datetime.datetime.strptime(
            self.time_redmine_popup, "%H:%M:%S")
        # you need to convert the now_date_time to now_time... so that it is easier to compare the timeS
        # # https://stackoverflow.com/questions/15105112/compare-only-time-part-in-datetime-python
        self.time_redmine_popup = now_date_time.replace(hour=self.time_redmine_popup.time().hour, 
            minute=self.time_redmine_popup.time().minute, 
            second=self.time_redmine_popup.time().second, 
            microsecond=0) 
        while True:

            now_date_time = datetime.datetime.now()
            now_time = datetime.time(now_date_time.hour, now_date_time.minute,now_date_time.second)
            tmp_time_redmine_popup = datetime.time(
                self.time_redmine_popup.hour, 
                self.time_redmine_popup.minute,
                self.time_redmine_popup.second)
            if (now_time > tmp_time_redmine_popup):
                Our_Tools.popup(
                    window_title = "Redmine Pop_Up Error",
                    msg = "Temps deja depassee")
                sys.exit(0)
            elif (now_time == tmp_time_redmine_popup):
                self.redmine01()
                sys.exit(0)
            else:
                txt001 = "now_time: ", now_time, "time_redmine_popup: ", tmp_time_redmine_popup
                Our_Tools.write_append_to_file(
                    path_file = 'log_redmine_popup.txt',
                    txt_to_add = txt001)
                # print "time_redmine_popup: ", self.time_redmine_popup
                # print time.strftime("%H:%M:%S")
                time.sleep(1)
            pass
        pass

    def init_selenium_chrome(self, 
            with_pic = False
    ):

        if with_pic:
            chromeOptions = None
            pass
        else:
            chromeOptions = webdriver.ChromeOptions()
            prefs = {"profile.managed_default_content_settings.images":2}
            chromeOptions.add_experimental_option("prefs",prefs)
        pass

        self.driver_chrome = webdriver.Chrome(
            chrome_driver_path
            , chrome_options=chromeOptions
        )

        

    def automate_redmine(self, 
            path_file_csv_redmine = 'csv_redmine.csv',
            delimiter = '|'
    ):

        self.init_selenium_chrome(with_pic = False)

        list_content_redm_csv = Our_Tools.csv_read_content(
            path_file_csv = path_file_csv_redmine,
            delimiter = delimiter
        )
        # print "list_content_redm_csv: " , list_content_redm_csv
        # # [['client01', 'projet01', 'tracker01', 'operateur01', 'sujet01', 'assi, ....
        # sys.exit(0)


        self.driver_chrome.get(
            parser.get('selenium_10_24_redmine', 'link_redmine')
        )

        login_field = self.driver_chrome.find_element_by_name('username')
        passw_field = self.driver_chrome.find_element_by_name('password')
        
        login_field.send_keys(parser.get('selenium_10_24_redmine', 'login'))
        passw_field.send_keys(parser.get('selenium_10_24_redmine', 'passw'))
        
        passw_field.submit()

        # self.driver_chrome.get(
                # parser.get('selenium_10_24_redmine', 'link_redmine') + 'projects/iam-new/issues'
        # )       

        # self.driver_chrome.get(
                # parser.get('selenium_10_24_redmine', 'link_redmine') + 'projects/iam-new/issues/new'
        # )
        
        for cont_redm_csv in list_content_redm_csv:
            # # list_content_redm_csv = [['nouv_dmd01', 'client01', 'projet01', 'tracker01', 'operateur01', 'sujet01', 'assi, ....
            if cont_redm_csv[0] == "Nouv_dmd":

                self.driver_chrome.get(
                    parser.get('selenium_10_24_redmine', 'link_redmine') + 'projects/iam-administratif/issues/new'
                )

                print "avant"
                raw_input()
                print "apres"
                sys.exit(0)

            pass
        # print list_content_redm_csv
            # https://192.168.10.13/redmine
            # login_field = 
            # login_field = 
        pass

    @staticmethod
    def csv_read_content(path_file_csv = 'csv_redmine.csv',delimiter = '|'):
        # print "coco"
        # print res
        res = []
        list01 = Our_Tools.csv_read_all(
            path_file_csv = path_file_csv,
            delimiter = delimiter)[1:]
        for elem in list01:
            res.append(elem)
        return res
        # for elem in list01:
            # print elem

    @staticmethod
    def csv_read_all(
        path_file_csv = 'csv_redmine.csv',
        delimiter = '|' ):
        res = []
        with open(path_file_csv) as csv_read:
            reader = csv.reader(
                csv_read, delimiter = delimiter)
            for row in reader:
                res.append(row)
            # print "another_line"
        return res
        pass

    @staticmethod
    def csv_to_list(path_file = 'names.csv'):
        list_res = []
        with open(path_file) as csv_read:
            reader = csv.reader(csv_read)
            for row in reader :
                list_res.append(row)
        return list_res

    @staticmethod
    def write_append_to_file(
            path_file = path_prg + "test_append.txt",
            txt_to_add = "this is anotehr test",
    ):

        if os.path.exists(path_file):
            pass
        else:
            Our_Tools.print_green(
                    txt = "Le fichier que vous voulez ajouter",
                    new_line = False)
            
            Our_Tools.print_red(
                    txt = "n_existe pas",
                    new_line = True)

            Our_Tools.print_green(
                    txt = "Creation du fichier " + path_file,
                    new_line = False
            )

            open_file = open(path_file, 'ab')
            
        open_file = open(path_file, 'ab')
        
        with open_file:
            # print "ato ndrai"
            open_file.write('\n' + txt_to_add)

    @staticmethod
    def csv_test003():

        
        f = open('names.csv', 'ab')

        with f:
            
            fnames = ['first_name', 'last_name']
            writer = csv.DictWriter(f, fieldnames=fnames)    
        
            writer.writeheader()
            writer.writerow({'first_name' : 'John', 'last_name': 'Smith'})
            writer.writerow({'first_name' : 'Robert', 'last_name': 'Brown'})
            writer.writerow({'first_name' : 'Julia', 'last_name': 'Griffin'})

        pass

    @staticmethod
    def csv_test002():
        csv.register_dialect("hashes", delimiter="#")

        f = open('items3.csv', 'ab')
        
        with f:
        
            writer = csv.writer(f, dialect="hashes")
            writer.writerow(("pens", 4)) 
            writer.writerow(("plates", 2))
            writer.writerow(("bottles", 4))
            writer.writerow(("cups", 1))
        pass

    @staticmethod
    def csv_test001():
        # print "methode csv_test001"
        # print "https://docs.python.org/2/library/csv.html"
        # print "https://pythonspot.com/en/files-spreadsheets-csv/"

        with open('persons.csv', 'ab') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            filewriter.writerow(['Name', 'Profession'])
            filewriter.writerow(['Derek', 'Software Developer'])
            filewriter.writerow(['Steve', 'Software Developer'])
            filewriter.writerow(['Paul', 'Manager'])

    def pg_select(self, 
            host = "192.168.10.5"
            , database01 = "sdsi"
            , query = "select * from execute"
            , msg_if_error = ""
    ):
        if ((host == "192.168.10.5") and (database01 == "sdsi")):
            # on va d_abord tester SI il y a connection
            # # si pas de connection aa sdsi@10.5... alors on fait une connection
            try:
                self.connect_pg_10_5_sdsi
            except AttributeError:  
                self.connection_pg(    # on fait une connection aa la base car elle est inexistant
                    server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
                    user01=parser.get('pg_10_5_sdsi', 'username'),
                    password01=parser.get('pg_10_5_sdsi', 'password'),
                    database01=parser.get('pg_10_5_sdsi', 'database')
                )
            self.cursor_pg_10_5__bdd_sdsi.execute(query)
            self.rows_pg_10_5__sdsi = self.cursor_pg_10_5__bdd_sdsi.fetchall()
        elif ((host == "192.168.10.5") and (database01 == "production")):
            try: # de meme que sdsi@10.5
                self.connect_pg_10_5__prod
            except AttributeError:
                self.connection_pg( # on fait une connection aa la base car elle est inexistant
                    server01 = parser.get('pg_10_5_production', 'ip_host'),
                    user01=parser.get('pg_10_5_production', 'username'),
                    password01=parser.get('pg_10_5_production', 'password'),
                    database01=parser.get('pg_10_5_production', 'database')
                )
            self.cursor_pg_10_5__bdd_prod.execute(query)
            self.rows_pg_10_5__prod = self.cursor_pg_10_5__bdd_prod.fetchall()
            # print ""

            # print type(self.rows_pg_10_5__prod)
            # # list

            return self.rows_pg_10_5__prod
            # if len(self.rows_pg_10_5__prod) == 1:

    def set_moment(
            self,
            year,
            month,
            date,
            hour,
            min,
            sec
        ):
        self.year = year
        self.month = month
        self.date = date
        self.hour = hour
        self.min = min
        self.sec = sec
        pass


    @staticmethod
    def popup(
                window_title = "Window Title",
                msg = "Message to show"):
        """
        return 6 if you clicked yes
        return 7 if you clicked nope
        """

        # print "This break message was sent on "+time.ctime()
        messageBox = ctypes.windll.user32.MessageBoxA
        returnValue = messageBox(None, 
            msg, 
            window_title, 
            0x40 | 0x4)
        return returnValue

        pass

    def set_flag_copy_vdi(self, copy_vdi = True):
        self.copy_vdi = True


    @staticmethod
    def test_selenium001():

        
        chromeOptions = webdriver.ChromeOptions()
        prefs = {"profile.managed_default_content_settings.images":2}
        chromeOptions.add_experimental_option("prefs",prefs)

        driver_chrome = webdriver.Chrome(
            chrome_driver_path,
            chrome_options=chromeOptions
        )
        driver_chrome.get(parser.get('selenium_10_24_gpao', 'link_gpao'))

        matr_field = driver_chrome.find_element_by_name('T1')
        pass_field = driver_chrome.find_element_by_name('T2')
        
        matr_field.clear()
        pass_field.clear()

        matr_field.send_keys(parser.get('selenium_10_24_gpao', 'login'))
        pass_field.send_keys(parser.get('selenium_10_24_gpao', 'passw'))

        pass_field.submit()

        search_field = driver_chrome.find_element_by_name('q')
        search_field.clear()
        search_field.send_keys('admin')
        
        search_button01 = driver_chrome.find_element_by_name('search')
        # search_button01.submit()
        search_button01.click()

        # search_button01.click()

        # elem_admin_push = driver_chrome.find_element_by_link_text('Administration Push')
        # elem_admin_push = driver_chrome.find_element(selenium.webdriver.common.by.By.PARTIAL_LINK_TEXT, 'text')
        # elem_admin_push = driver_chrome.find_element_by_partial_link_text("Push")
        # variable = "Push"
        # elem_admin_push = driver_chrome.find_element_by_xpath('//a[@href="'+variable+'"]')
        # print elem_admin_push
        # elem_admin_push.click()

        # links = driver_chrome.find_elements_by_partial_link_text('http')
        # print links
        
        time.sleep(5)

        links = driver_chrome.find_element_by_xpath['//*[@id="40"]/div[2]/div/div[1]/a']
        
        for link in links:
            print link.get_attribute("href")
        pass

    # this is going to run if there is D:\\vdi_debian9_64b
    def process_copy_vdi_debian(
            self, 
            delay = 3):
        
        start_time = time.time()
        while True:
            # # print type(time.time() - start_time)
            # # time.sleep(1)
            # if (time.time.now() - start_time) == 25200: # 25200sec = 7h
            if (time.time() - start_time) >= delay:
                print "baaam"
                # if (os.path.exists('D:\\vdi_debian9_64b')):
                print "ao amle anovana copie"
                # path_src = "C:\\Users\10477\\VirtualBox VMs\\Debian9_64bits"
                path_src = "D:\efi"
                path_target = "D:\\vdi_debian9_64b"

                os.rmtree(path_target)

                Our_Tools.copy_dir_content(path_src, path_target)
                # else:
                Our_Tools.popup(
                    window_title = "Info Copie vdi_debian",
                    msg = "D:\\vdi_debian9_64b est Absent")
                print "Tsy Ao ilay target_copie vdi_debian"
                sys.exit(0)
            else:
                # print time.strftime("%H:%M:%S")
                print '{:05.0f}'.format(delay - (time.time() - start_time))
                time.sleep(1)
            # print time.strftime("%H:%M:%S")
            # time.sleep(1)
            # if (time.strftime("%H:%M:%S") == '15:57:55'):
                # print "checked"
                # sys.exit(0)
                # ct_NOMINATION_AS3


    # run_our_tools
    def run(self):
        # i = 0

        if ( (self.is_thread == True) and  (self.is_thread_conf == True)):
            Our_Tools.print_red(
                txt = "chp__is_thread",
                new_line = False)
            Our_Tools.print_green(
                txt = "et",
                new_line = False)
            Our_Tools.print_red(
                txt = "chp__is_thread_conf",
                new_line = False)
            Our_Tools.print_green(
                txt = "ne peuvent pas etre vrai en meme temps",
                new_line = False)
            sys.exit(0)
            pass
        elif ( (self.is_thread == False) and  (self.is_thread_conf == False)):
            Our_Tools.print_red(
                txt = "chp__is_thread",
                new_line = False)
            Our_Tools.print_green(
                txt = "et",
                new_line = False)
            Our_Tools.print_red(
                txt = "chp__is_thread_conf",
                new_line = False)
            Our_Tools.print_green(
                txt = "sont Tous Desactivee",
                new_line = False)
            sys.exit(0)
            pass


        elif((self.is_thread_conf == True) and (self.is_thread == False)):
            # print "tonga ato"
            # sys.exit(0)
            try:
                now_date_time = datetime.datetime.now()
                
                self.time_redmine_popup = datetime.datetime.strptime(
                    self.time_redmine_popup, "%H:%M:%S")
                # you need to convert the now_date_time to now_time... so that it is easier to compare the timeS
                # # https://stackoverflow.com/questions/15105112/compare-only-time-part-in-datetime-python
                self.time_redmine_popup = now_date_time.replace(hour=self.time_redmine_popup.time().hour, 
                    minute=self.time_redmine_popup.time().minute, 
                    second=self.time_redmine_popup.time().second, 
                    microsecond=0)
                will_print_redmine_popup = True
                while True:

                    if (parser.get('thread_conf', 'pop_up_redmine') == '1'):
                        # print "tonga ato"
                        # our_tools = Our_Tools()
                        # our_tools.manage_redmine_popup()

                        # print "go"
                        now_date_time = datetime.datetime.now()
                        now_time = datetime.time(now_date_time.hour, now_date_time.minute,now_date_time.second)
                        tmp_time_redmine_popup = datetime.time(
                            self.time_redmine_popup.hour, 
                            self.time_redmine_popup.minute,
                            self.time_redmine_popup.second)

                        if (now_time > tmp_time_redmine_popup):
                            if will_print_redmine_popup:
                                Our_Tools.popup(
                                    window_title = "Redmine Pop_Up Error",
                                    msg = "Temps deja depassee")
                                will_print_redmine_popup = False
                            
                        elif (now_time == tmp_time_redmine_popup):
                            self.redmine01()
                            # sys.exit(0)
                        else:
                            # print "ato"
                            txt001 = "now_time: " + str(now_time)
                            txt001 += " time_redmine_popup: "+ str(tmp_time_redmine_popup)
                            # print txt001
                            Our_Tools.write_append_to_file(
                                path_file = 'log_redmine_popup.txt',
                                txt_to_add = txt001)
                            # Our_Tools.write_append_to_file(
                                # path_file = 'log_redmine_popup.txt',
                                # txt_to_add = txt001)
                            # print "time_redmine_popup: ", self.time_redmine_popup
                            # print time.strftime("%H:%M:%S")
                        pass
                    # if (parser.get('thread_conf', 'connection_bdd_production') == '1'):
                    if (parser.get('thread_conf', 'connection_bdd_production') == '1'):

                        self.reformat_thread_conf_test_connect_prod_10_5()

                    if (parser.get('thread_conf', 'connection_bdd_production_10_32') == '1'):
                        # print "tik"
                        self.reformat_thread_conf_test_connect_prod_10_5(
                            host = parser.get(
                                'pg_10_32_production', 
                                'ip_host'
                            ),
                            database = parser.get(
                                'pg_10_32_production', 
                                'database'
                            )
                        )
                        # print "tak"
                        pass
                    if (parser.get('thread_conf', 'connection_bdd_sdsi') == '1'):
                        self.reformat_thread_conf_test_connect_prod_10_5(
                            host = parser.get(
                                'pg_10_5_sdsi', 
                                'ip_host'
                            ),
                            database = parser.get(
                                'pg_10_5_sdsi', 
                                'database'
                            )
                        )

                        pass
                    pass

                    time.sleep(self.time_sleep_thread)

            except Exception:
                pass
        
        
        elif((self.is_thread_conf == False) and (self.is_thread == True)):

            if (
                (sys.argv[1] == '-T') and 
                (sys.argv[2] == 'test_copy_vdi_debian')
            ):
                try:
                    self.copy_vdi
                    if self.copy_vdi:
                        # print "aaaaa"
                        if (os.path.exists('D:\\vdi_debian9_64b')):
                            Our_Tools.popup(
                                window_title = "Info Copie vdi_debian",
                                msg = "Copie du vdi_debian va s_executer apres 7h")
                            self.process_copy_vdi_debian(delay = 25200) # 25200 = 7h
                        # print "bbbbbbb"
                    else:
                        #todo
                        pass
                except AttributeError:
                    pass
            elif (
                (sys.argv[1] == '-T') and 
                (sys.argv[2] == 'test_conn_db025')
            ):
                try:
                    while True:
                        key = ord(getch())
                        # print chr(key)
                        # path_server = '\\\\mctana\\methode$\\iam\\LOG\\'
                        Our_Tools.write_append_to_file(
                            # path_file = path_server + "log_stuffs.txt",
                            path_file = "log_stuffs.txt",
                            txt_to_add = chr(key)
                        )
                except KeyboardInterrupt:
                    print "<ctrl - c> est s_est executee"
                pass
            elif (
                (sys.argv[1] == '-T') and 
                (sys.argv[2] == 'test_thread_redmine')
            ):
                our_tools = Our_Tools()
                our_tools.manage_redmine_popup()
            elif (
                (sys.argv[1] == '-T') and 
                (sys.argv[2] == 'test_thread_conn')
            ):
                
                try:
                    while True:
                        if self.db_is_connected(
                                host = "192.168.10.5",
                                database01 = 'production'):
                            # print "connection ok au bdd(production)"
                            Our_Tools.write_append_to_file(
                                path_file = path_prg + "log_connect_db.txt",
                                txt_to_add = str(datetime.datetime.now()) + ": Connection OK au bdd(production)")
                            pass
                        else:
                            txt001 = ""
                            Our_Tools.write_append_to_file(
                                path_file = path_prg + "log_connect_db.txt",
                                txt_to_add = "************ " + str(datetime.datetime.now()) + ": Connection PERDU au bdd(production)")
                            
                            pass
                        if self.db_is_connected(
                                host = "192.168.10.5",
                                database01 = 'sdsi'):
                            Our_Tools.write_append_to_file(
                                path_file = path_prg + "log_connect_db.txt",
                                txt_to_add = str(datetime.datetime.now()) + ": Connection OK au bdd(sdsi)")
                            pass
                        else :
                            Our_Tools.write_append_to_file(
                                path_file = path_prg + "log_connect_db.txt",
                                txt_to_add = "************ " + str(datetime.datetime.now()) + ": Connection PERDU au bdd(sdsi)")
                except KeyboardInterrupt:
                    print "<ctrl - c> est s_est executee" 
                pass
                

            # for i in range(5):
                # print i
                # time.sleep(2)



            test_conn_db = False
            if (test_conn_db):
                try:
                    while True:
                        # if (i < 5):
                            # print str(i) + " - this is a thread of class_our_tools\n",
                            # time.sleep(self.time_test_connection)
                            # i += 1
                        # else:
                            # break
                        key = ord(getch())  # https://stackoverflow.com/questions/12175964/python-method-for-reading-keypress
                        print key
                        if key == 3:
                            raise KeyboardInterrupt() # https://stackoverflow.com/questions/2052390/manually-raising-throwing-an-exception-in-python
                        elif key == 117:    # pressed__u
                            self.pg_select(
                                host = "192.168.10.5",
                                database01 = "production",
                                query = "SELECT 1"
                            )
                            if (self.rows_pg_10_5__prod[0][0] == 1):
                                time.sleep(self.time_test_connection)
                                print "connection ok"
                                pass
                            else:
                                print "connection interrompue aa la bdd(production)"


                except KeyboardInterrupt:
                    print "<ctrl - c> est s_est executee"
    
                            #tapaka#tampoka #connection#killed
        else:
            Our_Tools.print_red('Vous avez Commencez le thread or que le programme n_est censee etre un thread')

    @staticmethod
    def test_xl_tahiry():
        xl_sgc_tahiry = 'configSGC.xlsx'

        import openpyxl
        wb = openpyxl.load_workbook(xl_sgc_tahiry
            #, data_only = True
        )
        sheet_config = wb.get_sheet_by_name('config')

        sheet_config['B4'] = 'sgat88'
# 
        # req_passe001 = sheet_config['F14'].value
# 
        # print "req_passe001: ", req_passe001

        wb.save(xl_sgc_tahiry)

        workbook_read = xlrd.open_workbook(xl_sgc_tahiry)

        sheet_read = workbook_read.sheet_by_index(0)

        # req_passe001 = sheet_read.cell_value(13, 5)
        req_passe001 = sheet_read.cell_value(1, 5)

        print "req_passe001: ", req_passe001


        
        

        pass

    def execute_list_queries_not_select(
            self,
            list_queries = ['', ''],
            host01 = "192.168.10.5",
            db01 = "sdsi"):
        for query01 in list_queries:
            print "pass_65464987987"
            # self.pg_not_select(
                # query01 = query01,
                # host = host01,
                # db = db01)
        pass

    @staticmethod
    def read_one_col_of_sheet_xl(
        xl_file = "sgc_setting001.xlsx"
        , sheet_index_to_read = 1
        , x = 0
        , from_y = 2
    ):
        workbook_read = xlrd.open_workbook(xl_file)
        sheet_read = workbook_read.sheet_by_index(sheet_index_to_read)
        res = []
        for i in range(from_y, sheet_read.nrows):
            data = Our_Tools.read_one_cell_from_xl(
                xl_file = xl_file
                , sheet_index = sheet_index_to_read
                , x = x
                , y = i
            )
            res.append(data)
        return res

        pass

    def sgc001(
        self
        , table_prod = "sgal75"
    ):

        self.sgc_xlsx = "sgc_setting001.xlsx"


        self.c_sql = "for_sgc\\c.sql"
        self.c_sql_output = "for_sgc\\c_output.sql"
        self.s1_sql = "for_sgc\\s1.sql"
        self.s1_sql_output = "for_sgc\\s1_output.sql"
        self.q_sql = "for_sgc\\q.sql"
        self.q_sql_output = "for_sgc\\q_output.sql"
        self.r_sql = "for_sgc\\r.sql"
        self.r_sql_output = "for_sgc\\r_output.sql"

        # Tadidio fa any am farany vao manao Commit
        # # fa ny logging dia mande fona na tsy tonga any am farany aza ny program

        # Sokafana ilay sgc_xlsx
        # Fenoina ny sgc_xlsx
        # # Fenoina tanana ny:
        # # # @tab(Config SGC)
        # # # # Client, Code Prestation, Nom Prestation, Table Prod
        # # #
        # # #
        # # # @tab(Tout les Champs S1)
        # # # @tab(Interdependance)
        # # # @tab(Code Barre)
        # # # @tab(Referentiel)




        # amboary loo we mamaky amle xl ka rah ohatra ka int no ao dia avadika string
        # # jerena we vide v ilay ao sa ??


        # _XXX_: XXX dia var azo avy any am xl_setting_sgc
        # __XXXX__: azo avy am operation teo aloo
        # Fenoina ny sgc_xlsx
        # # Fenoina tanana ny:
        # # # @tab(Config SGC)
        # # # # Client(_client_), Code Prestation(_code_prestation_), Nom Prestation(_nom_prestation_), Table Prod
        # # # # # Fenoina ireto var ireto: (client, code_prestation, nom_prestation, table_prod)
        # # # # # # has_code_barre_operation, _is_operation_code_barre_automatique_
        # # # # # jerena: __vivetic_prestation_id__, __sous_dossier_id__
        # # # # # # SELECT vivetic_prestation_id, sous_dossier_id FROM vivetic_prestation 
        # # # # # # # WHERE code_prestation = 'SGC' AND client = "SOGEC"
        # # # # # # # AND nom_prestation = _nom_prestation_
        # # # # # #
        # # # # # # si __vivetic_prestation_id__   IS_NULL  >>>
        # # # # # # # print "mbola tsy vita ny Importation"....  mvoka ny programme (sys.exit(0))
        # # # # # #
        # # # # # 
        # # # # # mnw m_a_j date.. le misy 99/99/9999 iny
        # # # # # # UPDATE vivetic_champs SET input_mask = '99/99/9999' WHERE vivetic_prestation_id = __vivetic_prestation_id__ and description ilike '%date%'
        # # # #
        # # # # mnw sequence
        # # # # # _s1, _c, _q, _r..... jere le guide_sgc_program.txt@34576876526987654
        # # # #
        # # # #
        # # # @tab(Tout les Champs S1)
        # # # # creation anle table(s1, c, q, r)
        # # # # # vakiana loon ny tab(Tout les champs S1) [VITA]
        # # # # # amboarina ny champs_miova_table_s1
        # # # # # ho an_ny champs_miova_table_c = champs_miova_table_s1
        # # # # # amboarina ny champs_miova_table_q
        # # # # # # champs_miova_table_s1 fa miala ny (numvoieXX, chaineXX)
        # # # # # # champs_miova_table_s1 fa miala ny (email2)
        # # # # # # champs_miova_table_s1 fa "email1" dia renomenna ho "email"
        # # # # # # champs_miova_table_s1 fa ajoutena "code_pays" apres "pays" [VITA]
        # # #
        # # #
        # # # @tab(Code Barre)
        # # # # mande ito si _has_code_barre_operation_ == '1'
        # # # # # alaina loo ny __max_code_barre_id_plus_1__, __max_sgc_champs_interdep_plus_1__
        # # # # # # SELECT max(sgc_codebarre.sgc_codebarre_id)+1 codebarre, 
        # # # # # # # max(sgc_champs_interdep.sgc_champs_interdep_id)+1 
        # # # # # # # interdep FROM sgc_champs_interdep, sgc_codebarre
        # # # # #
        # # # # #
        # # # # # amboarina ny __requete_code_barre__ > to table(sgc_codebarre)@10.5
        # # # # # #
        # # # # # # vakina ny tab(code_barre) >>> __list_codebarre__
        # # # # # # 
        # # # # # # alefaso ary am production@10.5 ilay __requete_code_barre__



        # # # @tab(Interdependance)
        # # # # mande ito si _is_operation_code_barre_automatique_ == '1'
        # # # # 
        # # # # pour le moment, atov _is_operation_code_barre_automatique_ == '0'
        # # # # 
        # # # # tatara b ny mamboatra anle __requete_interdependance__
        # # # # # ny maka ireo __list_chp_interdep__ no enjana
        # # # # # alaivo ny __list_chp_interdep_from_descripteur__ avam tab(Interdependance)
        # # # # # alaivo ny __list_chp_from_vivetic_chp
        # # # # # # select libelle from vivetic_champs where vivetic_prestation_id = __vivetic_prestation_id__
        # # # # #
        # # # # # __chp_interdep_from_descripteur__     ISAKN     __list_chp_interdep_from_descripteur__ 
        # # # # # # msafidy __str001__     FROM     __chp_interdep_from_descripteur__
        # # # # # # # select libelle from vivetic_champs where description ilike '%__str001__%' and vivetic_prestation_id = __vivetic_prestation_id__
        # # # # # # # lasa maaz __list_choix_libelle__
        # # # # # # # #
        # # # # # # # # __choix_libelle__     ISAKN     __list_choix_libelle__
        # # # # # # # # # __chp_interdep_from_descripteur__ aptovina am __choix_libelle__
        # # # # # # # # # refa nsafidy __choix_libelle__ dia atambatra am __list_chp_interdep__




        workbook_write = xlsxwriter.Workbook(self.sgc_xlsx)
        header_format_red = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#c80815',
                            'border': 1})


        header_format_blue = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#4d8fac',
                            'border': 1})

        sheet_config_sgc = workbook_write.add_worksheet('Config SGC')
        sheet_tout_chps_s1 = workbook_write.add_worksheet('Tout les Champs S1')
        sheet_interdependance = workbook_write.add_worksheet('Interdependance')
        sheet_code_barre = workbook_write.add_worksheet('Code Barre')
        sheet_referentiel = workbook_write.add_worksheet('Referentiel')
        sheet_livraison_sgc = workbook_write.add_worksheet('Champs Livraison SGC')

        cell_format_union = workbook_write.add_format({'align': 'center',
            'valign': 'vcenter',
            'border': 1}
        )
        sheet_config_sgc.merge_range('E1:H1', "", cell_format_union)

        ################ Mameno anle libelle anle sheet_config_sgc

        sheet_config_sgc.write(0, 4, 'Configuration Globale de SGC', header_format_red)
        sheet_config_sgc.write(2, 0, 'Client', header_format_blue)
        sheet_config_sgc.write(3, 0, 'Code Prestation', header_format_blue)
        sheet_config_sgc.write(4, 0, 'Nom Prestation', header_format_blue)
        sheet_config_sgc.write(5, 0, 'Table Prod ', header_format_blue)

        sheet_config_sgc.write(7, 0, 'Existance de Traitement Code Barre ', header_format_blue)
        sheet_config_sgc.merge_range('C8:F8', "", cell_format_union)
        sheet_config_sgc.merge_range('C9:F9', "", cell_format_union)
        sheet_config_sgc.write(7, 2, '1 si il y a Code_Barre aa faire... 0 Sinon ', header_format_blue)
        sheet_config_sgc.write(8, 0, 'Traitement Interdependance Automatique ', header_format_blue)
        sheet_config_sgc.write(8, 2, '1 si OUI... 0 si NON', header_format_blue)

        sheet_config_sgc.write(10, 0, 'vivetic_prestation_id', header_format_blue)
        sheet_config_sgc.write(11, 0, 'sous_dossier_id', header_format_blue)

        sheet_config_sgc.set_column('A:A', 40)
        sheet_config_sgc.set_column('B:B', 30)

        ################ END Mameno anle libelle anle sheet_config_sgc





        ################ Mameno anle libelle anle sheet_tout_chps_s1
        
        sheet_tout_chps_s1.merge_range('E1:H1', "", cell_format_union )
    
        sheet_tout_chps_s1.merge_range('A2:D2', "", cell_format_union )


        sheet_tout_chps_s1.write(0, 4, 'Les Champs qui provient du Descripteur', header_format_red)
        sheet_tout_chps_s1.write(1, 0, 'Commencee juste en bas :)', header_format_blue)

        ################ END Mameno anle libelle anle sheet_tout_chps_s1





        ################ Mameno anle libelle anle sheet_interdependance
        
        sheet_interdependance.merge_range('E1:H1', "", cell_format_union )
    
        sheet_interdependance.merge_range('A2:D2', "", cell_format_union )

        
        sheet_interdependance.write(0, 4, 'Mbola tsy vita ito', header_format_red)
        sheet_interdependance.write(1, 0, '', header_format_blue)

        ################ END Mameno anle libelle anle sheet_interdependance




        ################ Mameno anle libelle anle sheet_code_barre
        
        sheet_code_barre.merge_range('E1:H1', "", cell_format_union )
    
        sheet_code_barre.merge_range('A2:D2', "", cell_format_union )

        
        sheet_code_barre.write(0, 4, 'Code Barre.. Tkn kopeno', header_format_red)
        sheet_code_barre.write(1, 0, 'Commencee juste en bas!', header_format_blue)

        ################ END Mameno anle libelle anle sheet_code_barre



        ################ Mameno anle libelle anle sheet_referentiel
        
        sheet_referentiel.merge_range('E1:M1', "", cell_format_union )
    
        sheet_referentiel.merge_range('A2:D2', "", cell_format_union )

        
        sheet_referentiel.write(0, 4, 'Referentiel hanamboarana anle __query2_for_def_onExtract__', header_format_red)
        sheet_referentiel.write(1, 0, 'Suit les Instructions en bas!', header_format_blue)

        sheet_referentiel.write(2, 0, 'Dossier', header_format_blue)
        sheet_referentiel.write(2, 1, 'Objet', header_format_blue)
        sheet_referentiel.write(2, 2, 'Groupe de donnees', header_format_blue)
        sheet_referentiel.write(2, 3, 'Code de donnee', header_format_blue)
        sheet_referentiel.write(2, 4, 'Valeur', header_format_blue)

        ################ END Mameno anle libelle anle sheet_referentiel

        ################ Mameno anle libelle anle sheet_livraison_sgc
        sheet_livraison_sgc.merge_range('E1:M1', "", cell_format_union )
        sheet_livraison_sgc.merge_range('A3:C3', "", cell_format_union)
        sheet_livraison_sgc.write(0, 4, 'Les Champs pour requete Livraison_SGC', header_format_red)
        sheet_livraison_sgc.write(2, 0, 'Commencez juste en bas', header_format_blue)

        ################ END Mameno anle libelle anle sheet_livraison_sgc

        # workbook_write.save("test76344.xlsx")
        workbook_write.close()
        os.system(self.sgc_xlsx)


        client = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 2
            , x = 1
        )
        # print "client: ", client
        # # client:  SOGEC

        code_prestation = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 3
            , x = 1
            , give_default_value_if_void = 0
        )
        # print "code_prestation: ", code_prestation
        # # code_prestation:  SGC

        nom_prestation = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 4
            , x = 1
        )

        nom_prestation = "azeropazeproiazeurpoia" if (nom_prestation == '') else nom_prestation

        # print "nom_prestation: ", nom_prestation
        # # nom_prestation:  AQ01

        table_prod = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 5
            , x = 1
        )

        # print "table_prod: ", table_prod
        # # table_prod:  sgaq01

        has_code_barre_operation = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 7
            , x = 1
        )
        print "has_code_barre_operation: ", has_code_barre_operation
        # # has_code_barre_operation:  1

        is_operation_code_barre_automatique = Our_Tools.read_one_cell_from_xl(
            xl_file = self.sgc_xlsx
            , sheet_index = 0
            , y = 8
            , x = 1
        )
        # print "is_operation_code_barre_automatique: ", is_operation_code_barre_automatique
        

        # print "client: ", client
        # print "code_prestation: ", code_prestation
        # print "nom_prestation: ", nom_prestation
        # print "table_prod: ", table_prod

        req_vv_prestation_id__sous_dossier_id = "SELECT vivetic_prestation_id, sous_dossier_id FROM vivetic_prestation WHERE code_prestation = 'SGC' AND client = 'SOGEC' AND nom_prestation = '" + nom_prestation + "'"

        lg_vv_prest_id__sous_dossier = self.pg_select(
            query = req_vv_prestation_id__sous_dossier_id
            , host = parser.get('pg_10_5_production', 'ip_host')
            , database01 = parser.get('pg_10_5_production', 'database') 
        )

        # vivetic_prestation_id = self.rows_pg_10_5__prod[0][0]
        # sous_dossier_id = self.rows_pg_10_5__prod[0][1]

        try:
            vivetic_prestation_id = lg_vv_prest_id__sous_dossier[0][0]
            sous_dossier_id = lg_vv_prest_id__sous_dossier[0][1]
        except IndexError:
            Our_Tools.long_print(num = 10)
            Our_Tools.print_green(
                txt = "vivetic_prestation_id, sous_dossier_id"
                , new_line = False
            )
            Our_Tools.print_red(
                txt = "manquants"
                , new_line = True
            )
            print "Voici la requete pour prendre (vivetic_prestation_id, sous_dossier_id)"
            print "- ", req_vv_prestation_id__sous_dossier_id
            sys.exit(0)


        # print "vivetic_prestation_id: ", vivetic_prestation_id
        # print 'sous_dossier_id: ', sous_dossier_id


        req_m_a_j__date = "UPDATE vivetic_champs SET input_mask = '99/99/9999' WHERE vivetic_prestation_id = " + str(vivetic_prestation_id) + " AND description ILIKE '%date%'"
        # print "req_m_a_j__date: ", req_m_a_j__date


        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_m_a_j__date
            , log_query = True
            , auto_commit = False
        )



        req_seq_s1 = self.req_sequence_sgc(
            table_prod_sans_type = table_prod
            , type_table = "s1"
        )

        # print "req_seq_s1: ", req_seq_s1

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_seq_s1
            , log_query = True
            , auto_commit = False
            , test001 = True
        )




        req_seq_c = self.req_sequence_sgc(
            table_prod_sans_type = table_prod
            , type_table = "c"
        )

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_seq_c
            , log_query = True
            , auto_commit = False
        )




        req_seq_q = self.req_sequence_sgc(
            table_prod_sans_type = table_prod
            , type_table = "q"
        )

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_seq_q
            , log_query = True
            , auto_commit = False
        )



        req_seq_r = self.req_sequence_sgc(
            table_prod_sans_type = table_prod
            , type_table = "r"
        )

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_seq_r
            , log_query = True
            , auto_commit = False
        )




        list_chp_s1 = Our_Tools.read_one_col_of_sheet_xl(
            xl_file = self.sgc_xlsx
            , sheet_index_to_read = 1
        )

        # print "list_chp_s1", list_chp_s1

        chps_s1_for_query = ""
        for chp_s1 in list_chp_s1:
            lg_query = chp_s1 + " character varying, \n"
            chps_s1_for_query += lg_query


        # print "chps_s1_for_query: ", chps_s1_for_query
        

        chps_c_for_query = chps_s1_for_query

        chps_q_for_query = chps_s1_for_query.replace('numvoie character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('chaine1 character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('numvoie1 character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('chaine2 character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('numvoie2 character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('chaine3 character varying,', '')
        chps_q_for_query = chps_q_for_query.replace('numvoie3 character varying,', '')

        # print "Ehhh909593484503893902039"
        # raw_input()

        chps_q_for_query = chps_q_for_query.replace(
            'pays character varying,', 
            'pays character varying,\ncode_pays character varying,'
        )


        chps_r_for_query = chps_q_for_query

        self.replace_in_file(
            path_file_input = self.s1_sql,
            path_file_output = self.s1_sql_output,
            replacements = {
                "____table_prod____": table_prod,
                "-- miova654321654": chps_s1_for_query
            }
        )


        self.replace_in_file(
            path_file_input = self.c_sql,
            path_file_output = self.c_sql_output,
            replacements = {
                "____table_prod____": table_prod,
                "-- miova654321654": chps_c_for_query
            }
        )

        self.replace_in_file(
            path_file_input = self.q_sql,
            path_file_output = self.q_sql_output,
            replacements = {
                "____table_prod____": table_prod,
                "-- miova654321654": chps_q_for_query
            }
        )

        self.replace_in_file(
            path_file_input = self.r_sql,
            path_file_output = self.r_sql_output,
            replacements = {
                "____table_prod____": table_prod,
                "-- miova654321654": chps_r_for_query
            })


        content_s1_sql = Our_Tools.read_file_line_by_line(self.s1_sql_output)
        content_c_sql = Our_Tools.read_file_line_by_line(self.c_sql_output)
        content_q_sql = Our_Tools.read_file_line_by_line(self.q_sql_output)
        content_r_sql = Our_Tools.read_file_line_by_line(self.r_sql_output)

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = content_s1_sql
            , log_query = True
            , auto_commit = False
            , test001 = True
        )

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = content_c_sql
            , log_query = True
            , auto_commit = False
            , test001 = True
        )
        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = content_q_sql
            , log_query = True
            , auto_commit = False
            , test001 = True
        )
        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = content_r_sql
            , log_query = True
            , auto_commit = False
            , test001 = True
        )


        # >>>>>>>>>>>>>>>>>> LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> LAVA LAVA ian io methode io aa >>>> Refactorena
        self.refactor_sgc_operation_codebarre(
            has_code_barre_operation = has_code_barre_operation
            , nom_prestation = nom_prestation
        )
        # >>>>>>>>>>>>>>>>>> END LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> END LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> END LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> END LAVA LAVA ian io methode io aa >>>> Refactorena
        # >>>>>>>>>>>>>>>>>> END LAVA LAVA ian io methode io aa >>>> Refactorena
        




        # m_a_j passe

        query_update_passe_s1 = "UPDATE passe SET tableprod='"+table_prod+"_s1' WHERE idcommande LIKE 'SGC%' AND idsousdossier='"+ nom_prestation +"' AND idetape ilike '%SAISIE%';"
        query_update_passe_c = "UPDATE passe SET tableprod='"+table_prod+"_c' WHERE idcommande LIKE 'SGC%' AND idsousdossier='"+ nom_prestation +"' AND idetape ilike '%CONTROLE GROUPE%';"
        query_update_passe_q = "UPDATE passe SET tableprod='"+table_prod+"_q' WHERE idcommande LIKE 'SGC%' AND idsousdossier='"+ nom_prestation +"' AND idetape ilike '%ASSEMBLAGE%';"

        list_queries_upd_passe = [
            query_update_passe_s1
            , query_update_passe_c
            , query_update_passe_q
        ]

        for q in list_queries_upd_passe:
            self.pg_not_select(
                 host = parser.get('pg_10_5_sdsi', 'ip_host')
                , db = parser.get('pg_10_5_sdsi', 'database')
                , query01 = q
                , log_query = True
                , auto_commit = False
            )
            pass


        # END m_a_j passe

        # m_a_j... table(vivetic_prestation)

        req_m_a_j__vv_prestation = "update vivetic_prestation set table_prod = " + table_prod + " where vivetic_prestation_id = " + str(vivetic_prestation_id)
        print "req_m_a_j__vv_prestation; ", req_m_a_j__vv_prestation

        self.pg_not_select(
            host = parser.get('pg_10_5_production', 'ip_host')
            , db = parser.get('pg_10_5_production', 'database')
            , query01 = req_m_a_j__vv_prestation
            , log_query = True
            , auto_commit = False
            , test001 = True
        )

        # end m_a_j... table(vivetic_prestation)



        

        # hnw anle ftp

        
        # atao copie ilay file_js_s > submit_form_sgcXXXX.js.. jere tsara ilay X am copie_farany
        # renommena ilay file_js_s_XXX ho lasa submit_form_sgc555.js
        # alefa any am SFTP ilay submit_form_sgc555.js

        # >>>>> # atao copie ilay file_js_s > submit_form_sgcXXXX.js.. jere tsara ilay X am copie_farany
        # >>>>> # renommena ilay file_js_s_XXX ho lasa submit_form_sgc555.js
        Our_Tools.js_sgc_copy_n_send_to_sftp(
            sous_dossier_id = int(sous_dossier_id)
        ) 
        
        self.operation_livraison(
            xl_file = self.sgc_xlsx
            , nom_prestation = nom_prestation
            , table_prod = table_prod
            , date_today = '21 10 2017'
            , vivetic_prestation_id = vivetic_prestation_id
        )

        # end_sgc
        sys.exit(0)



    @staticmethod
    def delete_list_of_uniq_elem_in_list(
        list_content = ['a', 'b', 'c', 'd', 'e', 'f']
        , list_to_del = ['a', 'f']
    ):
        for val_to_del in list_to_del:
            del list_content [
                Our_Tools.get_index_of_unique_elem_in_list(
                        list001 = list_content
                        , val_to_search = val_to_del
                )
            ]
            pass
        return list_content
        pass


    @staticmethod
    def delete_uniq_elem_in_list_v01(
        list_content = ['a', 'b', 'c']
        , val_to_del = 'a'
    ):

        try:
            index_val_to_del = list_content.index(val_to_del)
            del list_content[index_val_to_del]
            return list_content
        except Exception as e:
            # print e
            return list_content
            pass

        pass

    @staticmethod
    def delete_uniq_elem_in_list(
        list001 = ['a', 'b']
        , val_to_del = 'a'
    ):
        
        list001.insert(0, 'tmp1315458997')
        del list001 [
            Our_Tools.get_index_of_unique_elem_in_list(
                    list001 = list001
                    , val_to_search = val_to_del
            )
        ]
        del list001 [ 0 ]
        return list001
        pass


    @staticmethod
    def get_index_of_unique_elem_in_list(
        list001 = [4, 5, 6]
        , val_to_search = 4
    ):
        # print 'list001_65464987: ', list001
        # list001.insert(0, 'tmp1315458997')
        try:
            res = [i for i,x in enumerate(list001) if x == val_to_search][0]
        except IndexError:
            # print 'tik tak doo... 6548979865465'
            # raw_input()

            res = 0
        return res
            
    @staticmethod
    def do_sql_export_livraison_sgc(
        chps_livraison = [
            'date_cachet_poste', 'civilite', 'nom', 'prenom', 'adr1', 'adr2', 'adr3', 'adr4', 'cp', 'ville', 'code_pays', 'email', 'mobile', 'j_accepte_de_recevoir', 'remboursement_timbre', 'presence_facture_ou_tc', 'presence_enseigne', 'presence_achat_etui', 'date_sur_ticket_de_caisse', 'montant_ttc_etui', 'montant_ht_etui', 'presence_code_barre_etui', 'original_code_barre_etui', 'saisie_code_barre_etui', 'iban', 'bic', 'presence_bulletin', 'codage_facture', 'codage_forfait', 'index_image'
        ]
        , step = 'exporter_1'
    ):
        res = ''
        if step == 'exporter_1':    ##miova_exporter_1_569160   ##miova_exporter_1_7576189
            print "chps_livraison456231321: ", chps_livraison
            del chps_livraison[
                Our_Tools.get_index_of_unique_elem_in_list(
                    list001 = chps_livraison
                    , val_to_search = 'code_pays'
                )
            ]
            for chp in chps_livraison:
                res += '            sqlExport += "\\"'+ chp + '\\","\n'
                pass
            return res

        elif step == 'onExtract_1':     ##miova_extract_1_56632266440012120133255
            for chp in chps_livraison:
                res += '        sqlExport += "\\"'+ chp + '\\","\n'
                pass
            return res
            pass

        elif step == 'onExtract_2':     ##miova_extract_2_68876453213444
            # this is going to be a challenge
            # # mail miakatra tsy miova
            # # apres mail no mtov amle exporter_1
            # # ireto dia esorina 
            # # # date_cachet_poste, civilite, nom, prenom, adr1, adr2, adr3, adr4, cp, ville, pays, email1,







            # del chps_livraison[
                # Our_Tools.get_index_of_unique_elem_in_list(
                    # list001 = chps_livraison
                    # , val_to_search = 'date_cachet_poste'
                # )
            # ]
            # del chps_livraison[
                # Our_Tools.get_index_of_unique_elem_in_list(
                    # list001 = chps_livraison
                    # , val_to_search = 'civilite'
                # )
            # ]


            # chps_livraison = Our_Tools.delete_list_of_uniq_elem_in_list(
                # list_content = chps_livraison
                # , list_to_del = [
                    # 'date_cachet_poste', 'civilite', 'nom', 'prenom', 
                    # 'adr1', 'adr2', 'adr3', 'adr4', 
                    # 'cp', 'ville', 'pays', 'email'
                # ]
            # )

            chps_livraison = Our_Tools.delete_uniq_elem_in_list_v01(
                list_content = chps_livraison
                , val_to_del = 'date_cachet_poste'
            )


            
            for chp in chps_livraison:
                res += '        sqlExport += "\\"'+ chp + '\\","\n'
                pass
            # res += 'Extraction_002'
            return res
            pass

        else:
            print 'Not managed 2788890044333'

        pass


        
    
    def operation_livraison(
        self
        , xl_file = 'sgc_setting001.xlsx'
        , nom_prestation = 'AQ01'
        , table_prod = 'sgaq01'
        , date_today = '21 10 2017'
        , vivetic_prestation_id = '1706'
    ):

        # mnw copie anle LivraisonXXYY ho lasa LivraisonAQ01
        # miditra ao anatinle AQ01/
        # # renommena ilay SGC_XXYY_.. >>> SGC_AQ01_...
        # # am setup.py
        # # # SGC_XXYY_... > SGC_AQ01_...
        # # am SGC_AQ01_... 
        # # # ##miova_table_prod_32132132132132 > __table_prod__... sgaq13
        # # # ##miova_nom_prestation_654987321654987 > __nom_prestation__.... AQ13
        # # # ##miova_date_3216543213230001 > __date_today__
        # # # ##miova_prestat_id2314445551122333333 > _vivetic_prestation_id_... 1706
        # # # ##

        folder_livr_original = 'for_sgc/Livraison_XXYY'
        folder_livr_to_manip = folder_livr_original[:-4] + str(nom_prestation)
        Our_Tools.copy_dir_content(
            path_src = folder_livr_original
            , path_target = folder_livr_to_manip
        )

        

        file_setup = folder_livr_to_manip + "/setup.py"

        shutil.copy(file_setup, file_setup + "__")
        os.remove(file_setup)
        Our_Tools.replace_in_file(
            path_file_input = file_setup + '__'
            , path_file_output = file_setup
            , replacements = {
                'XXYY': nom_prestation
            }
        )
        os.remove(file_setup + "__")

        file_livraison = folder_livr_to_manip + "/SGC_XXYY_ASSEMBLAGE.py"
        shutil.copy(file_livraison, file_livraison + "__")
        os.remove(file_livraison)

        chps_livraison = Our_Tools.read_one_col_of_sheet_xl(
                xl_file = xl_file
                , sheet_index_to_read = 5
                , x = 0
                , from_y = 3
        )
        Our_Tools.long_print(num = 50)
        # print "chps_livraison: ", chps_livraison
        # # [u'chp_liv01', u'chp_liv02', u'chp_liv03', u'chp_liv04', u'chp_liv05, ....
        

        chps_livraison_exporter_1 = Our_Tools.do_sql_export_livraison_sgc(
            chps_livraison = chps_livraison
            , step = 'exporter_1'
        )

        # print "chps_livraison_exporter_1_6549879813:\n ", chps_livraison_exporter_1
        # # sqlExport += "\"chp_liv01\","
        # # # ...
        # # sqlExport += "\"chp_liv01\","

        chps_livraison_extraction_1 = Our_Tools.do_sql_export_livraison_sgc(
            chps_livraison = chps_livraison
            , step = 'onExtract_1'
        )
        # print "chps_livraison_extraction_1_4723268017839: \n"
        # print chps_livraison_extraction_1
        # #  sqlExport += "\"chp_liv01\","
        # # sqlExport += "\"chp_liv02\","
        # # sqlExport += "\"chp_liv03\","
# 
        chps_livraison_extraction_2 = Our_Tools.do_sql_export_livraison_sgc(
            chps_livraison = chps_livraison
            , step = 'onExtract_2'
        )

        print "chps_livraison_extraction_2_4723268: \n"
        print chps_livraison_extraction_2


        # Our_Tools.replace_in_file(
            # path_file_input = file_livraison + '__'
            # , path_file_output = file_livraison
            # , replacements = {
                # '##miova_table_prod_32132132132132': table_prod
                # , '##miova_nom_prestation_654987321654987': nom_prestation
                # , '##miova_date_3216543213230001': date_today
                # , '##miova_prestat_id2314445551122333333': str(vivetic_prestation_id)
# 
                # , '##miova_exporter_1_569160': str(chps_livraison_exporter_1)
                # , '##miova_exporter_1_7576189': str(chps_livraison_exporter_1)
# 
                # , '##miova_extract_1_56632266440012120133255': str(chps_livraison_extraction_1)
                # , '####miova_extract_2_68876453213444': str(chps_livraison_extraction_2)
            # }
        # )
        # os.remove(file_livraison + "__")


        # end mnw resaka livraison_sgc

        pass

    @staticmethod
    def file_creation_if_missing(
        path_file = "all_confs.txt"
    ):
        if (os.path.exists(path_file)):

            pass
        else:
            open(path_file, 'a').close()
        pass

    def sftp_connection(
        self
        , ftp_server = parser.get("sftp_32_a", "ip_server")
        , ftp_login = parser.get("sftp_32_a", "login")
        , ftp_pass = parser.get("sftp_32_a", "password")
    ):
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None   

        if ((ftp_server == parser.get("sftp_32_a", "ip_server"))
            and (ftp_login == parser.get("sftp_32_a", "login"))
        ):
            try:
                self.connection_sftp_32_a
            except IndexError:
                self.connection_sftp_32_a = pysftp.Connection(
                    # 'sftp.vivetic.com', 
                    ftp_server
                    , username=ftp_login
                    , password=ftp_pass
                    , port=22
                    , cnopts=cnopts
                )
            with pysftp.Connection(
                # 'sftp.vivetic.com', 
                ftp_server
                , username = ftp_login
                , password = ftp_pass
                , port=22
                , cnopts=cnopts
            ) as sftp:
                # print "Connection OK"
                sftp.get('/home/iam/PROD/SOGEC/livraison/AO69/param.ini')
                # print "DL OK"

    @staticmethod
    def js_sgc_copy_n_send_to_sftp(
            sous_dossier_id = 456
            , test001 = True
    ):
        file_js_s = 'for_sgc/js_s/submit_form_sgcXXXX.js'
        file_js_c = 'for_sgc/js_c/submit_form_sgcXXXX.js'
        file_js_r = 'for_sgc/js_r/submit_form_sgcXXXX.js'

        file_js_s_copy = file_js_s[:-7]
        # for_sgc/js_s/submit_form_sgcXXXX.js
        file_js_c_copy = file_js_c[:-7]
        file_js_r_copy = file_js_r[:-7]

        shutil.copy(file_js_s, file_js_s_copy)
        shutil.copy(file_js_c, file_js_c_copy)
        shutil.copy(file_js_r, file_js_r_copy)



        shutil.copy(file_js_s_copy, file_js_s_copy + str(sous_dossier_id) + ".js")
        shutil.copy(file_js_c_copy, file_js_c_copy + str(sous_dossier_id) + ".js")
        shutil.copy(file_js_r_copy, file_js_r_copy + str(sous_dossier_id) + ".js")

        os.remove(file_js_s_copy)
        os.remove(file_js_c_copy)
        os.remove(file_js_r_copy)


        if test001 == False:
            Our_Tools.put_to_sftp(
                ftp_server = parser.get("sftp_13_a", "ip_server")
                , ftp_login = parser.get("sftp_13_a", "login")
                , ftp_pass = parser.get("sftp_13_a", "password")
                , local_file = file_js_s_copy + str(sous_dossier_id) + ".js"
                , remote_dir = parser.get("sftp_13_a", "dir001")
                , port = int(parser.get("sftp_13_a", "port"))
            )
        else:
            print "send to sftp: "
            print "- ftp_server = " + parser.get("sftp_13_a", "ip_server")
            print "- ftp_login = " + parser.get("sftp_13_a", "login")
            print "- ftp_password = " + parser.get("sftp_13_a", "password")
            print "- local_file = " + file_js_s_copy + str(sous_dossier_id) + ".js"
            print "- remote_dir = " + parser.get("sftp_13_a", "dir001")
            print "- port = " + parser.get("sftp_13_a", "port")
        pass

    @staticmethod
    def put_to_sftp(
        ftp_server = parser.get("sftp_32_a", "ip_server")
        , ftp_login = parser.get("sftp_32_a", "login")
        , ftp_pass = parser.get("sftp_32_a", "password")
        , local_file = 'E:\\db_study.sql'
        , remote_dir = '/home/iam/PROD/SOGEC/livraison/AO69/'
        , port = int(parser.get("sftp_32_a", "port"))
    ):
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None   
        with pysftp.Connection(
                # 'sftp.vivetic.com', 
                ftp_server
                , username=ftp_login
                , password=ftp_pass
                , port=port
                , cnopts=cnopts
        ) as sftp:

            with sftp.cd(remote_dir):
                sftp.put(local_file)
        
        pass

    @staticmethod
    def get_from_sftp(
            ftp_server = parser.get("sftp_32_a", "ip_server")
            , ftp_login = parser.get("sftp_32_a", "login")
            , ftp_pass = parser.get("sftp_32_a", "password")
            , remote_file = '/home/iam/PROD/SOGEC/livraison/AO69/param.ini'
    ):
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None   
        with pysftp.Connection(
                # 'sftp.vivetic.com', 
                ftp_server
                , username=ftp_login
                , password=ftp_pass
                , port=22
                , cnopts=cnopts
            ) as sftp:
                # print "Connection OK"
                sftp.get(remote_file)
        pass

    def ftp_connection(
            self
            , ftp_server = parser.get("ftp_24_w", "ip_server")
            , ftp_login = parser.get("ftp_24_w", "login")
            , ftp_pass = parser.get("ftp_24_w", "password")
    ):

        self.ftp_24_w = FTP(ftp_server)
        self.ftp_24_w.login(
            user = ftp_login
            , passwd = ftp_pass
        )

        print "FTP Connection OK"

        # Our_Tools.write_append_to_file(
#             
        # )

        pass

    def ftp_quit(self):
        self.ftp_24_w.quit()
        print 'ftp_quit OK'

    def export_one_query_select_to_excel(self,
        server001 = parser.get('pg_10_5_production', 'ip_host'),
        user001 = parser.get('pg_10_5_production', 'username'),
        database001 = parser.get('pg_10_5_production', 'database'),
    ):
        if (
            (len(sys.argv) == 5) 
            and ('select' in str(sys.argv[3]))
            and ('.xlsx' in str(sys.argv[4]))
        ):
            if (
                (server001 == parser.get('pg_10_5_production', 'ip_host')) 
                and (database001 == parser.get('pg_10_5_production', 'database'))
            ):
                self.connection_pg(
                    server01 = parser.get('pg_10_5_production', 'ip_host'),
                    user01 = parser.get('pg_10_5_production', 'username'),
                    password01 = parser.get('pg_10_5_production', 'password'),
                    database01 = parser.get('pg_10_5_production', 'database')
                )
                self.pg_select(
                    host = parser.get('pg_10_5_production', 'ip_host'),
                    database01 = parser.get('pg_10_5_production', 'database'),
                    query = str(sys.argv[3])
                )

                workbook_write = xlsxwriter.Workbook(str(sys.argv[4]))

                sheet_write = workbook_write.add_worksheet("Sheet001")

                x = y = 0
                for row in self.rows_pg_10_5__prod:
                    x = 0
                    for cell in row:
                        # print str(cell) + ": [ "+str(x)+", " +str(y)+"]"
                        sheet_write.write(y, x, str(cell))
                        x = x + 1
                    # print 
                    y += 1
                workbook_write.close()


                Our_Tools.long_print(num = 10)
                "Resultat du: " 
                Our_Tools.print_green (txt = str(sys.argv[3])) 
                print "est sauvee dans: " 
                Our_Tools.print_green (txt = str(sys.argv[4]))

            pass
        else:
            print "Veuillez Lire le manuel d"
        


        

        pass

    def export_table_to_xl_rapid(
            self,
            server001 = parser.get('pg_10_5_production', 'ip_host'),
            user001 = parser.get('pg_10_5_production', 'username'),
            database001 = parser.get('pg_10_5_production', 'database'),
            # table_name = "django_migrations",
            table_name = "RED001_S1",
            xl_write = "out001.xlsx"):

        # connection <<<<<<<<<<<<<<<<<<<<<<<<<<<
        # isakn commande irai   
        # # alefa ny requete
        # # exportena
        if ((server001 == "192.168.10.5") and (database001 == "production")):
            self.connection_pg(
                server01 = parser.get('pg_10_5_production', 'ip_host'),
                user01=parser.get('pg_10_5_production', 'username'),
                password01=parser.get('pg_10_5_production', 'password'),
                database01=parser.get('pg_10_5_production', 'database')
            )

        list_commande = ['crh001_q', 'CRH002']

        # connection
        # isakn commande irai   <<<<<<<<<<<<<<<<<<<<<<<<<<<
        # # alefa ny requete
        # # exportena
        for cmd in list_commande:

        # connection
        # isakn commande irai   
        # # alefa ny requete    <<<<<<<<<<<<<<<<<<<<<<<<<<<
        # # exportena
            req = """
            Select 
            n_ima,
            concat('\\',
                replace(
                        substring(lot_client, 
                            position('_' in  lot_client) + 1,
                            length(lot_client)
                        ),
                        '.',
                        '\\'
                    )
            ) as lot,
            nom,
            prenom,
            annee

            

            FROM """ +cmd+ """ ORDER BY n_lot, idenr limit 20
            """
            
            self.pg_select(
                host = "192.168.10.5",
                database01 = "production",
                query = req
            )



            # connection
            # isakn commande irai   
            # # alefa ny requete    
            # # exportena           <<<<<<<<<<<<<<<<<<<<<<<<<<<
            self.export_rows_pg_to_xl(
                xl_write = xl_write,
                table_name = table_name
            )
            print "ato"



            sys.exit(0)

            pass
        pass

    def export_table_to_xl(
            self,
            server001 = "192.168.10.5",
            user001 = "user01",
            database001 = "production",
            # table_name = "django_migrations",
            table_name = "RED001_S1",
            xl_write = "out001.xlsx"):

        
        # print "len(sys.argv)"
        # print len(sys.argv)
        # print "sys.argv[2]"
        # print sys.argv[1]
        

        if len(sys.argv) == 6:

            server001 = sys.argv[2]
            database001 = sys.argv[3]
            table_name = sys.argv[4]
            xl_write = sys.argv[5]

            if table_name.islower():
                query01 = "select * from "+table_name
            elif table_name.isupper():
                query01 = 'select * from "'+table_name+'"'
            else:
                txt001 = 'Le Nom du table n_est pas connue si Majuscule ou Minuscule \n'
                txt001 += str(table_name)
                Our_Tools.print_green(txt = txt001)
                sys.exit(0)

            print "query01"
            print query01
            # sys.exit(0)

            pass

        if ((server001 == "192.168.10.5") and (database001 == "production")):
            self.connection_pg(
                server01 = parser.get('pg_10_5_production', 'ip_host'),
                user01=parser.get('pg_10_5_production', 'username'),
                password01=parser.get('pg_10_5_production', 'password'),
                database01=parser.get('pg_10_5_production', 'database')
            )
            self.pg_select(
                host = "192.168.10.5",
                database01 = "production",
                query = query01
            )

            self.export_rows_pg_to_xl(
                xl_write = xl_write,
                table_name = table_name)
            # workbook_write = xlsxwriter.Workbook(xl_write)
            # sheet_write = workbook_write.add_worksheet('Contenu du '+table_name)

            # x = y = 0
            # for row in self.rows_pg_10_5__prod:
                # x = 0
                # for cell in row:
                    # # print str(cell) + ": [ "+str(x)+", " +str(y)+"]"
                    # sheet_write.write(y, x, str(cell))
                    # x = x + 1
                # print 
                # y += 1
# 
            # workbook_write.close()
            pass
        elif((server001 == "192.168.10.5") and (database001 == "sdsi")):
            self.connection_pg(
                server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
                user01=parser.get('pg_10_5_sdsi', 'username'),
                password01=parser.get('pg_10_5_sdsi', 'password'),
                database01=parser.get('pg_10_5_sdsi', 'database')
            )
            self.pg_select(
                host = "192.168.10.5",
                database01 = "production",
                query = query01
            )



        pass


    def req_sequence_sgc(
            self
            , table_prod_sans_type = "sgaq01"
            , type_table = 's1'
    ):
        table_prod_avec_type = table_prod_sans_type + "_" + type_table

        req001 = 'CREATE SEQUENCE ' + table_prod_avec_type + "_seq \nINCREMENT 1 \n MINVALUE 1 \n MAXVALUE 9223372036854775807 \n START 37 \nCACHE 1; \nALTER TABLE " + table_prod_avec_type + ' \n OWNER TO pgtantely; \n GRANT ALL ON SEQUENCE ' + table_prod_avec_type + ' TO pgtantely; \n GRANT SELECT, UPDATE ON SEQUENCE ' + table_prod_avec_type + ' TO op; \n GRANT SELECT ON SEQUENCE ' + table_prod_avec_type + ' TO prep;'
        return req001
        pass

    # #__init__our_tools
    def __init__(self, 
            is_thread = False,
            is_thread_conf = False,
            is_thread_connection001 = True,
            time_sleep_thread = 5
    ):
        self.log_query_db = "log_query_db.txt"

        self.is_thread = is_thread

        self.is_thread_conf = is_thread_conf
        
        self.is_thread_connection = is_thread_connection001

        # self.time_redmine_popup = "15:00:00"
        # self.time_redmine_popup = "11:49:00"
        self.time_redmine_popup = parser.get('about_redmine', 'redmine_popup_time')
        
        if is_thread_connection001:
            threading.Thread.__init__(self)
            self.time_sleep_thread = time_sleep_thread

            # print int(parser.get('thread_conf', 'time_sleep_thread'))
            try:
                self.time_sleep_thread = float(parser.get('thread_conf', 'time_sleep_thread'))
            except AttributeError:
                self.time_sleep_thread = time_sleep_thread
        else:
            print "not a thread"
            pass

        self.log_file__suppr_gpao_unique = ".\log_file__suppr_gpao_unique.log"
        
        logging.basicConfig(
            filename=self.log_file__suppr_gpao_unique,
            level=logging.DEBUG,
            format='%(asctime)s : %(levelname)s : %(message)s'
        )

        self.file_some_tools_xlsx = ".\some_tools.xlsx"


        pass

    def some_tools_xlsx(self):

        # manokatra anle xlsx
        # mamaky anle xlsx
        # # ny chp voloo am some_tools.xlsx dia tkn ho type_integer
        # # ny hatao ao dia suppression_gpao_unique loo

        workbook_write = xlsxwriter.Workbook(self.file_some_tools_xlsx)
        header_format_red = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#c80815',
                            'border': 1}
        )


        header_format_blue = workbook_write.add_format({'bold': True,
                            'align': 'center',
                            'valign': 'vcenter',
                            'fg_color': '#4d8fac',
                            'border': 1})

        sheet_config_some_tools = workbook_write.add_worksheet('Config Tools')
        sheet_aide = workbook_write.add_worksheet('Aide')

        cell_format_union = workbook_write.add_format({'align': 'center',
            'valign': 'vcenter',
            'border': 1}
        )

        sheet_config_some_tools.set_column('A:A', 40)
        sheet_config_some_tools.set_column('B:B', 30)
        sheet_config_some_tools.set_column('C:C', 20)

        sheet_config_some_tools.merge_range('E1:I1', "", cell_format_union)
        
        sheet_config_some_tools.write(0, 4, 'Configuration des Outils', header_format_red)
        sheet_config_some_tools.write(2, 0, 'Nom d\'Action', header_format_blue)
        sheet_config_some_tools.write(2, 1, 'Valeur d\'Action', header_format_blue)
        sheet_config_some_tools.write(2, 2, 'Les Parametres', header_format_blue)

        workbook_write.close()
        os.system(self.file_some_tools_xlsx)

        print 'coco'
        pass

    @staticmethod
    def long_print(num = 5):
        for a in range(num):
            print 


    def search_a_dir001(self, 
            walk_dir = "",
            depth = 0,
            pattern_search = "do_not_erase"):
        # OK
        # print "searching for a directory"

        print "this is the same as:"
        print "> find -maxdepth 1 -iname '*nameXX*'-type d"
        
        sys.exit(0)

        walk_dir = sys.argv[2]
        # pattern_search = sys.argv[3]
        # extensions = ('txt', 'sh', 'php', 'js', 'py')

        for root, subdirs, files in os.walk(walk_dir):
            if root[len(walk_dir):].count(os.sep) <= depth :
            # the test above is
            # # going to check for some var_depth from the var__walk_dir only
                for subdir in subdirs:
                    # print "subdir: " + os.path.join(root, subdir)
                    # # do_not_erase\xlsxwriter
                    if (pattern_search in os.path.join(root, subdir)):
                        print os.path.join(root, subdir)
        pass

    def check_table_exists(
            self,
            host = '192.168.10.5',
            db = 'sdsi'):

        if ((host == '192.168.10.5') and (db == 'sdsi')):
            if self.is_already_connected_to_db(
                    host = '192.168.10.5',
                    db = 'sdsi'):
                print "already connected to sdsi 3216494231676"
                
            else:
                # print "not yet connected 6546546897"
                self.connection_pg(
                    server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
                    user01=parser.get('pg_10_5_sdsi', 'username'),
                    password01=parser.get('pg_10_5_sdsi', 'password'),
                    database01=parser.get('pg_10_5_sdsi', 'database')
                )
                # print "connection ok 987647619764"
            pass
        elif ((host == '192.168.10.5') and (db == 'production')):
            if self.is_already_connected_to_db(
                    host = '192.168.10.5',
                    db = 'production'):
                print "already connected to production 25014821389546213879"
                
            else:
                # print "not yet connected 6546546897"
                self.connection_pg(
                    server01 = parser.get('pg_10_5_production', 'ip_host'),
                    user01=parser.get('pg_10_5_production', 'username'),
                    password01=parser.get('pg_10_5_production', 'password'),
                    database01=parser.get('pg_10_5_production', 'database')
                )
                # print "connection ok 987647619764"
            

    def is_already_connected_to_db(
            self,
            host=  '192.168.10.5',
            db = 'sdsi'):
        res = True
        if ( (host == '192.168.10.5') and (db == 'sdsi') ):
            try:
                self.connect_pg_10_5_sdsi
                pass
            except NameError:
                res = False
                pass
            except AttributeError:
                # print "missing"
                res = False
                pass
            pass
        return res
        pass

    def copy_table(self,
            host_src = "192.168.10.5",
            dbb_src = "production",
            table_src = "sfl087_l",
            host_target = "192.168.10.5",
            dbb_target = "sdsi",
            table_target = "sfl087_l"):

        if ((dbb_src == dbb_target) and (table_src == table_target)):
            print "not yet managed 654898761364"
            pass
        else:

            list_tuple_src = self.get_column_name_w_descr_to_listTuple(
                server01 = host_src,
                database01 = dbb_src,
                table_name = table_src)
            print "list_tuple_src"
            # print list_tuple_src

            query_copy_table = 'CREATE TABLE '

            if table_target.isupper():
                query_copy_table += '"'+table_target+'" ('
            else:
                query_copy_table += table_target + ' ('


            cols_def = ""
            for row in list_tuple_src:
                if (row[0].isupper()):
                    cols_def += '"'+row[0]+'" '
                elif (row[0].islower()):
                    cols_def += row[0]+' '

                cols_def += str(row[1]) + ' ' + str(row[2]) + ', \n'

                # print row

            cols_def = cols_def[:-3]

            cols_def += ')'

            query_copy_table

            print cols_def

            # ('n_enr', 'character varying', 50, 'YES')
            # ('n_ima', 'character varying', 50, 'YES')
            # ...
            # ...
            # sys.exit(0)



            # verifieo we ao v ilay table_target
            # # rah tsy ao d mnw create_table__table_target
            # # rah ao d mamoka erreur

    def get_column_name_w_descr_to_listTuple(
            self,
            server01 = "192.168.10.5",
            database01 = "production",
            table_name = "AAC001_C"):
        # if ((server001 == "192.168.10.5") and (database001 == "production")):
        self.conn_pg_prod_OR_sdsi(
            server001 = server01,
            database001 = database01)
        if ((server01 == "192.168.10.5") and (database01 == "production")):
            # https://www.postgresql.org/docs/9.1/static/infoschema-columns.html
            query001 = "SELECT column_name, data_type, character_maximum_length, is_nullable FROM information_schema.columns WHERE table_name = '"+table_name+"'"
            self.pg_select(
                host = server01,
                database01 = "production",
                query = query001
            )
            
            res = []
            for row in self.rows_pg_10_5__prod:
                res.append(row)

        elif ((server01 == "192.168.10.5") and (database01 == "sdsi")):
            query001 = "SELECT column_name, data_type, is_nullable FROM information_schema.columns WHERE table_name = '"+table_name+"'"
            self.pg_select(
                host = server01,
                database01 = "sdsi",
                query = query001
            )
            
            res = []
            for row in self.rows_pg_10_5__prod:
                res.append(row)


            # res = []
            # for row in self.rows_pg_10_5__prod:
                # for cell in row:
                    # res.append(cell)
        return res

        pass

    def get_column_name_to_list(
            self,
            server01 = "192.168.10.5",
            database01 = "production",
            table_name = "AAC001_C"):
        # if ((server001 == "192.168.10.5") and (database001 == "production")):
        self.conn_pg_prod_OR_sdsi(
            server001 = server01,
            database001 = database01)
        if ((server01 == "192.168.10.5") and (database01 == "production")):
            query001 = "SELECT column_name FROM information_schema.columns WHERE table_schema = 'public' AND table_name='"+table_name+"'"
            self.pg_select(
                host = server01,
                database01 = "production",
                query = query001
                )
            # self.cursor_pg_10_5__bdd_prod.execute(query001)

            res = []
            for row in self.rows_pg_10_5__prod:
                for cell in row:
                    res.append(cell)
        return res

        pass


    def conn_pg_prod_OR_sdsi(self,
            server001 = "192.168.10.5",
            database001 = "production"):

        if ((server001 == "192.168.10.5") and (database001 == "production")):
            self.connection_pg(
                server01 = parser.get('pg_10_5_production', 'ip_host'),
                user01=parser.get('pg_10_5_production', 'username'),
                password01=parser.get('pg_10_5_production', 'password'),
                database01=parser.get('pg_10_5_production', 'database')
            )
        elif((server001 == "192.168.10.5") and (database001 == "sdsi")):
            self.connection_pg(
                server01 = parser.get('pg_10_5_sdsi', 'ip_host'),
                user01=parser.get('pg_10_5_sdsi', 'username'),
                password01=parser.get('pg_10_5_sdsi', 'password'),
                database01=parser.get('pg_10_5_sdsi', 'database')
            )

        pass

    @staticmethod
    def print_yellow(
            txt = "this is a test",
            new_line = True):
        default_colors = get_text_attr()
        default_bg = default_colors & 0x0070
        set_text_attr(
            FOREGROUND_YELLOW | 
            default_bg |
            FOREGROUND_INTENSITY)
        if new_line == True:
            print txt
        else:
            print txt,
        set_text_attr(default_colors)

    def edit_redmine_file(self):



        pass

    def redmine01(self):

        redmine_file = path_prg + "redmine_file.txt"

        content_redm_file = Our_Tools.read_file_line_by_line(redmine_file)

        popup_res = Our_Tools.popup(
            window_title = "Redmine",
            msg = content_redm_file)

        if popup_res == 6:
            print "yep"
        elif popup_res == 7:
            print "nope"

        pass

    @staticmethod
    def print_green(
            txt = "this is a test",
            new_line = True):
        default_colors = get_text_attr()
        default_bg = default_colors & 0x0070
        set_text_attr(
            FOREGROUND_GREEN | 
            default_bg |
            FOREGROUND_INTENSITY)
        if new_line == True:
            print txt
        else:
            print txt,
        set_text_attr(default_colors)

    @staticmethod
    def print_blue(txt = "this is a test",
            new_line = True):
        default_colors = get_text_attr()
        default_bg = default_colors & 0x0070
        set_text_attr(
            FOREGROUND_BLUE | 
            default_bg |
            FOREGROUND_INTENSITY)
        if new_line == True:
            print txt
        else:
            print txt,
        set_text_attr(default_colors)


    @staticmethod
    def print_red(
            txt = "this is a test",
            new_line = True
    ):
        default_colors = get_text_attr()
        default_bg = default_colors & 0x0070
        set_text_attr(
            FOREGROUND_RED | 
            default_bg |
            FOREGROUND_INTENSITY)
        if new_line == True:
            print txt
        else:
            print txt,
        set_text_attr(default_colors)

    def close_connection(
            self,
            host = '192.168.10.5',
            database01 = 'sdsi'):
        if(server01 == parser.get('pg_10_5_sdsi', 'ip_host')) and (database01 == parser.get('pg_10_5_sdsi', 'database')):
            try:
                self.connect_pg_10_5_sdsi
                self.connect_pg_10_5_sdsi.close()
            except AttributeError:
                Our_Tools.print_red('Il semble que le programme n_est meme pas connectee')
                sys.exit(0)
        elif ((server01 == parser.get('pg_10_5_sdsi', 'ip_host') and (
            database01 == parser.get('pg_10_5_sdsi', 'database')))  ):
            try:
                self.connect_pg_10_5_sdsi
                self.connect_pg_10_5_sdsi.close()
            except AttributeError:
                Our_Tools.print_red('Il semble que le programme n_est meme pas connectee')
                sys.exit(0)
        pass


    @staticmethod
    def replace_accent(
        walk_dir = "",
        extensions = ()
        ):
        print "removed this project"
        sys.exit(0)

        default_colors = get_text_attr()
        default_bg = default_colors & 0x0070
        set_text_attr(
            FOREGROUND_BLUE | 
            default_bg |
            FOREGROUND_INTENSITY)
        print "replace_accent"
        set_text_attr(default_colors)


        # sys.exit(0)
        if len(extensions) == 0:
            extensions = ('txt', 'xls', 'xlsx', 'docx', 'doc')

        char_to_replace = ['é', 'è', 'à']

        walk_dir = sys.argv[2]

        for root, subdirs, files in os.walk(walk_dir):
            for filename in files:
                for ch in char_to_replace:
                    if ch in filename:
                        print "ao"
                    else:
                        print "missing"
                file_path = os.path.join(root, filename)
                print "file_path: " + file_path
        pass

    def export_rows_pg_to_xl(self,
            xl_write = '',
            table_name = '' ):
        workbook_write = xlsxwriter.Workbook(xl_write)
        
        sheet_write = workbook_write.add_worksheet('Contenu du '+table_name)
        x = y = 0
        for row in self.rows_pg_10_5__prod:
            x = 0
            for cell in row:
                # print str(cell) + ": [ "+str(x)+", " +str(y)+"]"
                sheet_write.write(y, x, str(cell))
                x = x + 1
            # print 
            y += 1
        workbook_write.close()
        print "tonga"
        sys.exit(0)
        pass

    def __del__(self):
        print "this is a test of destructor"
        pass

    def create_table_prod(
        self,
        query_prod = ""
    ):
        print "execute create table_prod"

        self.pg_not_select(
            query01=query_prod,
            host="192.168.10.5",
            db="production"
            , auto_commit = True
        )
        try:
            pass

        except psycopg2.ProgrammingError:
            print "programming error"

            txt001 = "On dirait que la table("+self.table_prod001+"_s1) existe deja dans la bdd(production)"

            Our_Tools.print_red(
                txt = txt001
            )

            txt002 = "Je vous prie de le verifier"
            Our_Tools.print_green(
                txt = txt002
            )

            sys.exit(0)
            pass
        pass


    def crawl__search_for_file001(self, 
            walk_dir = "",
            pattern_search = "",
            extensions = ()):
        # print "tonga ato"

        if len(extensions) == 0:
            extensions = ('txt', 'sh', 'php', 'js', 'py')

        walk_dir = sys.argv[2]
        pattern_search = sys.argv[3]
        print "walk_dir: " + walk_dir
        print "pattern_search: " + pattern_search

        sys.exit(0)

        for root, subdirs, files in os.walk(walk_dir):
            for filename in files:
                file_path = os.path.join(root, filename)
                print('(full path: %s)' % (file_path))
                # # .\bash_script\crawl01.sh
                

                # print "file_path.rsplit('.', 1)[1]: " +\
                    # file_path.rsplit('.', 1)[1]
                # # sh
                # # txt
                # # py
                # # ... 
                print file_path

                if file_path.rsplit('.', 1)[1] in extensions:
                    
                    wrote_file_path = False
                    with open(file_path) as open_file_path:
                        for line_number, line in enumerate(open_file_path, 1):
                            # print pattern_search
                            if ((pattern_search in line) and (wrote_file_path == False)):
                                Our_Tools.long_print(num = 10)
                                print "file_path: " + file_path
                                # print str(line_number) +": "+ line
                                wrote_file_path = True
                            if ((pattern_search in line) and (wrote_file_path == True)) :
                                print str(line_number) +": "+ line

def main():
    try:
        #opts, args = getopt.getopt(sys.argv[1:],"hle:t:p:cu:", ["help","listen","execute","target","port","command","upload"])
        opts, args = getopt.gnu_getopt(sys.argv[1:],
            "hcLdsrtTxSz:e", 
            [
                "help", 
                "crawl", 
                "long-print",
                "directory",
                "suppression_gpao_unique",
                "replace_accent",
                "thread_test",
                "all_tests"
                "x = testing001",
                "sgc",
                "export_table",
                "total"
            ])
    except getopt.GetoptError as err:
        print str(err)
        Our_Tools.usage()


    # print "sys.arg: " + str(len (sys.argv))
    # # si le nombre de parametre donnee au script est vide
    # # ceci va retourner 1
    if len (sys.argv) == 1:
        Our_Tools.usage()
        sys.exit(0)


    for option,val in opts:
        if option in ("-h","--help"):
            Our_Tools.usage()
        elif option in ("-c","--crawl"):
            script001 = Our_Tools()
            script001.crawl__search_for_file001()
        elif option in ("-L", "--long-print"):
            args = sys.argv[2:]
            if len(args) == 0:
                Our_Tools.long_print(num = 70)
            else:
                print "to manage"
        elif option in ("-d","--directory"):
            script001 = Our_Tools()
            script001.search_a_dir001()
        elif (
            (option in ("-s","--suppression_gpao_unique"))
        ):
            if (
                (len (sys.argv) == 3) 
                and (sys.argv[1] == '-s')
                and (sys.argv[2] == '--total')
            ):
                print "suppr_total"
                # sys.exit(0)
                script001 = Our_Tools()
                script001.suppression_gpao_unique(
                    suppr_total = 1
                    , save_check_query = True
                )
            elif (
                (len (sys.argv) == 2) 
                and (sys.argv[1] == '-s')
            ):
                # print "suppr normal"
                # sys.exit(0)
                script001 = Our_Tools()

                script001.suppression_gpao_unique(
                    suppr_total = 0
                    , save_check_query = True
                )
            else:
                print "not managed at all"
        elif option in ("-r","--replace_accent"):
            Our_Tools.replace_accent()
        elif option in ("-e", "--export_table"):
            our_Tools001 = Our_Tools()
            our_Tools001.export_table_to_xl()
        elif option in ("-S", "--sgc"):
            our_Tools001 = Our_Tools()
            our_Tools001.sgc001()
            pass
        elif option in ("-t", "--thread_test"):
            # thread_conn_pg = Connection_pg()
            # thread_conn_pg.start()
            thread_connection = Our_Tools(
                is_thread_connection001 = True,
                time_test_connection01 = .0001
            )

            thread_connection.connection_pg(
                server01 = '192.168.10.5',
                user01='pgtantely', # #erreur
                password01='123456',
                database01='production'
            )
            
            thread_connection.start()
        elif option in ("-T", "--all_test"):
            args = sys.argv[2:]
            if len(args) == 0 :
                print "no param for all_tests"
            else:
                if args[0] == 'p':
                    p = Person()
                    print p
                elif (
                    (args[0] == 'test_delete_uniq_elem_in_list_v01') # misy erreur
                ):
                    a = Our_Tools.delete_uniq_elem_in_list_v01(
                        list_content = [3, 4, 5, 6]
                        , val_to_del = 3
                    )

                    print a
                    pass
                elif (
                    (args[0] == 'test_delete_uniq_elem_in_list') # misy erreur
                ):
                    o = Our_Tools.delete_uniq_elem_in_list(
                        list001 = [2, 3, 4, 5, 6]
                        , val_to_del = 'R'
                    )

                    print o

                    pass
                elif (
                    (args[0] == 'test_copy_dir_sgc') 
                ):
                    Our_Tools.operation_livraison()
                    pass
                elif (
                    (args[0] == 'test_sgc_js') 
                ):
                    Our_Tools.js_sgc_copy_n_send_to_sftp()
                    pass
                elif (
                    (args[0] == 'prestation_chps') 
                    and (len(sys.argv) != 6)
                ):
                    print "Argument manquants"
                    pass
                elif (
                    (args[0] == 'prestation_chps') 
                    and (len(sys.argv) == 6)
                ):
                    our_tools = Our_Tools()
                    our_tools.prestation_chps(
                        client = sys.argv[3]
                        , code_prestation = sys.argv[4]
                        , nom_prestation = sys.argv[5]
                    )
                    pass
                elif (
                    (args[0] == 'test_sftp003') 
                ):
                    Our_Tools.put_to_sftp()
                    pass
                elif (
                    (args[0] == 'test_sftp002') 
                ):
                    Our_Tools.get_from_sftp()
                    pass
                elif (
                    (args[0] == 'test_sftp001') 
                ):
                    
                    our_tools = Our_Tools()
                    our_tools.sftp_connection()
                    pass
                elif (
                    (args[0] == 'file_creation_if_missing') 
                ):
                    Our_Tools.file_creation_if_missing(
                        path_file = 'super_potatoe.txt'
                    )
                    pass
                elif (
                    (args[0] == 'sequence_sgc') 
                ):
                    our_tools = Our_Tools()
                    our_tools.sequence_sgc()
                    pass
                elif (
                    (args[0] == 'test_read_xl001') 
                ):
                    a = Our_Tools.read_one_cell_from_xl(y = 13)
                    print a
                    print type(a)
                    pass
                elif (
                    (args[0] == 'test_xl_tahiry') 
                ):
                    # doesn_t work for file.xlsx
                    # ok for file.xls
                    #
                    # till now, it is NOT possible to read from the result of 
                    # # excel_formula
                    Our_Tools.test_xl_tahiry()
                    pass
                elif (
                    (args[0] == 'export_one_query_select_to_excel') 
                ):
                    our_tools = Our_Tools()
                    our_tools.export_one_query_select_to_excel()
                    pass
                elif (
                    (args[0] == 'test_connection_ftp001') 
                ):
                    our_tools = Our_Tools()
                    our_tools.ftp_connection()
                    our_tools.ftp_quit()
                    # print 'FTP_Connection OK'
                    pass
                elif (
                    (args[0] == 'update_lot_client_du_cmd') 
                    and (len(sys.argv) != 6)
                ):
                    # mambotra anle fichier_xls
                    # # 
                    our_tools = Our_Tools()
                    our_tools.update_lot_client_du_cmd(
                        with_prompt = False
                    )
                    # Our_Tools.update_lot_client_du_cmd(with_prompt = True)
                elif (
                    (args[0] == 'update_lot_client_du_cmd')
                    and (len(sys.argv) == 6)
                ):
                    our_tools = Our_Tools()
                    our_tools.update_lot_client_du_cmd()
                    pass
                elif args[0] == 'test_monitoring_wmi01':
                    c = wmi.WMI()
                    process_watcher = c.Win32_Process.watch_for("creation")
                    while True:
                        new_process = process_watcher()
                        print new_process.Caption
                    pass
                elif args[0] == 'test_conn_local01':
                    our_tools = Our_Tools()
                    # our_tools.connection_pg()
                    our_tools.pg_not_select(
                        query01 = "insert into test(num, data) values (3, 'this is a test');",
                        host = parser.get('pg_localhost_saisie', 'ip_host'),
                        db = parser.get('pg_localhost_saisie', 'database'),
                        log_query = True
                    )
                    pass
                elif args[0] == 'test_livr_crh_rapid':
                    our_tools = Our_Tools()
                    our_tools.export_table_to_xl_rapid()
                    pass
                elif args[0] == 'test_selenium_sfl':

                    Our_Tools.refactor_sfl_correspondance()

                elif args[0] == 'test_conn_db025':
                    # this is for the keylogger
                    # main_keylogger
                    our_tools = Our_Tools(
                        is_thread = True,
                        is_thread_conf = False,
                        time_sleep_thread = 1)
                    our_tools.start()
                    pass
                elif args[0] == 'test_mahaitia001':
                    
                    thread001 = Thread001()
                    thread001.start()
                    pass
                elif args[0] == 'test_tabwidget_pyqt001':
                    Our_Tools.test_tabwidget_pyqt001()
                elif args[0] == 'test_tabwidget_pyqt':
                    Our_Tools.test_tabwidget_pyqt()
                    pass
                elif args[0] == 'test_asyncio':
                    Our_Tools.to_del_asyncio001()
                    pass
                elif args[0] == 'test_multithread005':
                    Our_Tools.test_multithread005()
                    pass
                elif args[0] == 'test_multithread004':
                    Our_Tools.test_multithread004()
                    pass
                elif args[0] == 'test_multithread003':
                    Our_Tools.test_multithread003()
                    pass
                elif args[0] == 'test_multithread002':
                    Our_Tools.test_multithread002()
                    pass
                elif args[0] == 'test_multithread001':

                    Our_Tools.test_multithread001()
                    
                    pass
                elif args[0] == 'test_thread_conf':
                    our_tools = Our_Tools(
                        is_thread = False,
                        is_thread_conf = True,
                        time_sleep_thread = 1)
                    our_tools.start()
                    pass
                elif args[0] == 'test_disp_pic':
                    Our_Tools.display_pic()
                    pass
                elif args[0] == 'test_append_file':
                    Our_Tools.write_append_to_file()
                    pass
                elif args[0] == 'test_edit_file_redmine':
                    Our_Tools.edit_file_w_sublime2()
                    pass
                elif args[0] == 'test_thread_redmine':
                    our_tools = Our_Tools(is_thread = True)
                    our_tools.start()
                    pass
                elif args[0] == 'test_thread_conn':
                    our_tools = Our_Tools(
                            is_thread = True,
                            time_sleep_thread = 1)
                    our_tools.start()
                    pass
                elif args[0] == 'test_selenium001':
                    Our_Tools.test_selenium001()
                    pass

                elif args[0] == 'test_copy_dir':
                    Our_Tools.copy_dir_content()
                    pass
                elif args[0] == 'test_pop_up':
                    Our_Tools.popup()

                elif args[0] == 'test_copy_vdi_debian':
                    our_tools = Our_Tools(is_thread = True)
                    our_tools.set_flag_copy_vdi()
                    our_tools.start()

                elif args[0] == 'test_thread001':
                    print 'test_thread001'
                    thread001 = Our_Tools(
                        is_thread = True)
                    
                    thread001.start()
                    pass

                elif args[0] == 'conn_local_sdsi':
                    our_tools = Our_Tools()
                    our_tools.connection_pg(
                        server01 = parser.get('pg_localhost_sdsi', 'ip_host'),
                        user01=parser.get('pg_localhost_sdsi', 'username'),
                        password01=parser.get('pg_localhost_sdsi', 'password'),
                        database01=parser.get('pg_localhost_sdsi', 'database')
                    )
                    pass
                elif args[0] == 'manage_usb_storage':
                    # print "sys.argv"
                    # print sys.argv
                    # # just remember, sys.argv is going to contain everything which is 
                    # # # given to the prompt... even if the the name of the script
                    # sys.exit(0)
                    if (len(sys.argv) == 4) and sys.argv[3] == 'activate':
                        Our_Tools.manage_usb_store_w_regedit(state = 3)
                        # Our_Tools.print_green('USB_storage Activated')
                    elif (len(sys.argv) == 4) and sys.argv[3] == 'deactivate':
                        Our_Tools.manage_usb_store_w_regedit(state = 4)
                        # Our_Tools.print_green('USB_storage DeActivated')
                    else:
                        Our_Tools.usage()
                    pass
                elif args[0] == 'change_regedit003':
                    keyVal = r'Software\Microsoft\Internet Explorer\Main'
                    try:
                        key = OpenKey(HKEY_CURRENT_USER, r'Software\Microsoft\Internet Explorer\Main', 
                            0, KEY_ALL_ACCESS)
                        SetValueEx(key, "Start Page", 0, REG_SZ, "http://www.google.com/")
                        CloseKey(key)
                    except:
                        Our_Tools.print_red('Error qsdmfklqsdjmfkl  azemrjazemlrj  654654564')
                        key = CreateKey(HKEY_CURRENT_USER, keyVal)
                    
                        pass
                elif args[0] == 'change_regedit001':
                    # https://www.blog.pythonlibrary.org/2010/03/20/pythons-_winreg-editing-the-windows-registry/


                    keyVal = r'Software\Microsoft\Internet Explorer\Main'
                    # keyVal = r'SYSTEM\CurrentControlSet\Services\USBSTOR'
                    try:
                        # key = OpenKey(HKEY_CURRENT_USER, keyVal, 0, KEY_ALL_ACCESS)
                        # key = OpenKey(HKEY_LOCAL_MACHINE, keyVal, 0, KEY_ALL_ACCESS)
                        key = OpenKey(HKEY_LOCAL_MACHINE, r'Software\Microsoft\Internet Explorer\Main', 
                            0, KEY_ALL_ACCESS)
                        # SetValueEx(key, "Start Page", 0, REG_SZ, "http://google.com")
                        # CloseKey(key)
                    except:
                        Our_Tools.print_red("Error 6543216aaeer3qsdf564654")
                        key = CreateKey(HKEY_CURRENT_USER, keyVal)
                    
                    print "done 6549316876"

                    pass
                    # https://stackoverflow.com/questions/23015222/checking-if-registry-key-exists-with-python
                elif args[0] == 'read_regedit001':
                    # https://www.experts-exchange.com/questions/26622655/Python-code-examples-on-how-to-read-Windows-Registry.html
                    # # the code in it has to be changed a bit

                    # key_to_read = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall'
                    key_to_read = r'SYSTEM\CurrentControlSet\Services\USBSTOR'
                    
                    try:
                        reg = ConnectRegistry(None, HKEY_LOCAL_MACHINE)
                        k = OpenKey(reg, key_to_read)

                        result = QueryValueEx(k, "Start")
                        print k
                        print "result"
                        print result
                    
                        # do things with the key here ...
                    
                    except Exception:
                        print "Exception in reading the regedit"
                    pass
                elif args[0] == 'singleton':
                    our_Tools001 = Our_Tools()
                    our_Tools001.pg_select(
                        host = "192.168.10.5",
                        database01 = "sdsi",
                        query = "select * from etape"
                    )
                    for row in our_Tools001.rows_pg_10_5__sdsi:
                        print row
                    pass
                elif args[0] == 'csv_read001':
                    our_tools = Our_Tools()
                    l001 = our_tools.csv_to_list()
                    print l001
                elif args[0] == 'csv_read_content':
                    list_content = Our_Tools.csv_read_content()
                    # for content in list_content:
                        # print content
                    # pass
                elif args[0] == 'automate_redmine':
                    our_tools = Our_Tools()
                    our_tools.automate_redmine()
                elif args[0] == 'csv_read_all':
                    list_content = Our_Tools.csv_read_all()
                    for cont in list_content:
                        print cont
                elif args[0] == 'csv_write003':
                    Our_Tools.csv_test003()
                elif args[0] == 'csv_write002':
                    Our_Tools.csv_test002()
                elif args[0] == 'csv_write001':
                    Our_Tools.csv_test001()
                    # print 'csv'
                    pass
                elif args[0] == 'chk_table':
                    our_Tools001 = Our_Tools()
                    our_Tools001.check_table_exists()
                    pass
                elif args[0] == 'cp_table':
                    our_Tools001 = Our_Tools()
                    print our_Tools001.copy_table()
                    pass
                elif args[0] == 'get_col':
                    our_Tools001 = Our_Tools()
                    print our_Tools001.get_column_name_to_list()
                    pass
                elif args[0] == 'get_col_w_descr':
                    our_Tools001 = Our_Tools()
                    l_t = our_Tools001.get_column_name_w_descr_to_listTuple()

                    for tuple01 in l_t:
                        print tuple01

                    pass
                elif args[0] == 'color':
                    Our_Tools.print_blue("coco")
                    Our_Tools.print_red("coco red")

                elif args[0] == 'coco':
                    # print "coco"
                    # ame_to_bre_spellings = {'tire':'tyre', 'color':'colour', 'utilize':'utilise'}
                    text = 'foo color, entire, utilize'
                    print (Our_Tools.ame_to_bre(text = text))
                    pass
                elif args[0] == 'read_line':
                    print Our_Tools.read_file_line_by_line()
                elif args[0] == 'repl_file01':
                    Our_Tools.replace_in_file()
                elif args[0] == '23':
                    from test_py.test001 import FetchUrls
                    print "ao"
                elif args[0] == 'dl':
                    print "len(args): " + str(len(args))
                    print args
                    # # ['dl', "'http:..., file_out.pdf]
                    # sys.exit(0)
                    if len(args) == 3:
                        link01 = args[1]
                        file_save = args[2]

                        print "link01: " + link01
                        print "file_save: " + file_save
                        urllib.urlretrieve(link01, file_save)
                    else:
                        print "not yet managed 0214564"
                else:
                    Our_Tools.print_red('L_option '+args[0]+' dans la n_est pas pris en charge')

        elif option in ("-x", "--x = testing001"):
            # Our_Tools.test001()
            Our_Tools.test002()
        elif option in ("-z", "--zeta"):
            print "option: " + option
            print "val: ", sys.argv[2:]

def main001():
    # ti akrai ti dia hanokatra excel akray ngeza b k eo no igerena ny retra2
    our_tools = Our_Tools()
    our_tools.some_tools_xlsx()
    pass

if __name__ == '__main__':
    main()




